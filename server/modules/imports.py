from __future__ import annotations

import csv
import html
import io
import json
import re
import tempfile
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from fastapi import APIRouter, File, Form, HTTPException, Query, Request, Response, UploadFile
from pydantic import BaseModel, Field

from .admin_roles import (
    effective_data_scopes_for_context,
    json_download_response,
    require_permission,
    role_codes_for_context,
    safe_download_stem,
)
from .auth import (
    AuthContext,
    area_allowed,
    data_scope_value_allowed,
    effective_areas,
    effective_data_scope_values,
    has_effective_area_scope,
    has_effective_data_scope,
    request_context,
)
from .business import (
    MAP_LAYER_PERMISSIONS,
    append_map_layer_audit_event,
    enrich_map_layer_record,
    find_layer_record_for_upsert,
    layer_records,
    map_layer_changed_fields,
    normalize_record,
    save_layer_records,
    upsert_import_batch_layer_mysql,
)
from .forest_blocks import (
    block_identities_by_codes,
    block_by_code,
    is_rights_archive_like_block,
    load_all_blocks,
    normalize_block,
    require_target_area_allowed,
    sanitize_block_for_ledger,
    save_blocks,
)
from .forest_scene_links import (
    load_scene_links,
    require_visible_scene,
    save_import_batch_scene_links_mysql,
    save_scene_link_mysql,
    save_scene_links,
    upsert_scene_link_postgis,
)
from .forest_rights import upsert_right_archives_from_blocks
from .database import (
    import_batches_json_path,
    load_json_records,
    mysql_connect,
    save_json_records,
    use_mysql,
    use_postgis,
)
from .settings import get_settings


router = APIRouter(prefix="/api", tags=["forest-imports"])

IMPORT_REPORTS: dict[str, dict[str, Any]] = {}
IMPORT_ERROR_FIELD = "_importErrors"
IMPORT_FOREST_BLOCKS_VIEW_PERMISSION = "imports.forestBlocks.view"
IMPORT_FOREST_BLOCKS_PERMISSION = "imports.forestBlocks.manage"
IMPORT_FOREST_BLOCKS_CREATE_PERMISSION = "imports.forestBlocks.create"
IMPORT_FOREST_BLOCKS_REVIEW_PERMISSION = "imports.forestBlocks.review"
IMPORT_FOREST_BLOCKS_QUALITY_PERMISSION = "imports.forestBlocks.quality"
IMPORT_FOREST_BLOCKS_ACCEPTANCE_PERMISSION = "imports.forestBlocks.acceptance"
IMPORT_FOREST_BLOCKS_ROLLBACK_PERMISSION = "imports.forestBlocks.rollback"
IMPORT_FOREST_BLOCKS_DELETE_PERMISSION = "imports.forestBlocks.delete"
IMPORT_FOREST_BLOCKS_RESTORE_PERMISSION = "imports.forestBlocks.restore"
IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION = "imports.forestBlocks.export"
IMPORT_SCENE_LAYER_LINK_PERMISSION = "imports.sceneLayers.link"
IMPORT_RELATION_EVENT_SAMPLE_LIMIT = 100
QUALITY_ISSUE_STATUSES = {"open", "investigating", "resolved", "ignored"}
IMPORT_ACCEPTANCE_STATUSES = {"pending", "accepted", "needs_correction", "rejected"}
IMPORT_DELIVERY_PACKAGE_STATUSES = {
    "awaiting_review",
    "awaiting_acceptance",
    "awaiting_scene_link",
    "awaiting_publish",
    "awaiting_delivery",
    "blocked",
    "ready",
}
IMPORT_DELIVERY_STATUSES = {"pending", "partial", "needs_correction", "rejected", "delivered"}
IMPORT_OPERATION_QUEUE_STAGES = [
    {
        "key": "awaiting_review",
        "label": "待审核",
        "description": "导入批次需要业务审核后才能进入验收和影像发布。",
        "tone": "warning",
        "primaryAction": "review",
        "primaryActionLabel": "提交审核",
        "requiredPermission": IMPORT_FOREST_BLOCKS_REVIEW_PERMISSION,
        "href": "admin-imports.html?deliveryPackageStatus=awaiting_review",
    },
    {
        "key": "awaiting_acceptance",
        "label": "待验收",
        "description": "审核已通过，需要记录图档一致和成果交付验收意见。",
        "tone": "review",
        "primaryAction": "acceptance",
        "primaryActionLabel": "记录验收",
        "requiredPermission": IMPORT_FOREST_BLOCKS_ACCEPTANCE_PERMISSION,
        "href": "admin-imports.html?deliveryPackageStatus=awaiting_acceptance",
    },
    {
        "key": "awaiting_scene_link",
        "label": "待挂影像",
        "description": "批次已验收，但还没有挂接影像或图层。",
        "tone": "warning",
        "primaryAction": "link-scene-layer",
        "primaryActionLabel": "关联影像图层",
        "requiredPermission": IMPORT_SCENE_LAYER_LINK_PERMISSION,
        "href": "admin-imports.html?deliveryPackageStatus=awaiting_scene_link",
    },
    {
        "key": "awaiting_publish",
        "label": "待发布",
        "description": "影像已关联，需要完成图层发布后进入交付。",
        "tone": "warning",
        "primaryAction": "publish-layer",
        "primaryActionLabel": "发布图层",
        "requiredPermission": IMPORT_SCENE_LAYER_LINK_PERMISSION,
        "requiredAllPermissions": ["map.layers.publish"],
        "requiredAnyPermissions": [["map.layers.create", "map.layers.update"]],
        "href": "admin-imports.html?deliveryPackageStatus=awaiting_publish",
    },
    {
        "key": "awaiting_delivery",
        "label": "待交付",
        "description": "图层已发布，需要影像目录记录交付结果和回执。",
        "tone": "review",
        "primaryAction": "record-delivery",
        "primaryActionLabel": "记录影像交付",
        "requiredPermission": "imagery.scenes.delivery",
        "href": "admin-imports.html?deliveryPackageStatus=awaiting_delivery",
    },
    {
        "key": "blocked",
        "label": "阻断",
        "description": "质量、发布风险或影像缺失造成闭环阻断，需要先处理问题。",
        "tone": "danger",
        "primaryAction": "resolve-quality",
        "primaryActionLabel": "处理阻断",
        "requiredPermission": IMPORT_FOREST_BLOCKS_QUALITY_PERMISSION,
        "href": "admin-imports.html?deliveryPackageStatus=blocked",
    },
    {
        "key": "ready",
        "label": "可交付",
        "description": "成果、验收、图层和影像交付均已完成，可导出交付材料。",
        "tone": "ready",
        "primaryAction": "export-delivery-package",
        "primaryActionLabel": "导出交付材料",
        "requiredPermission": IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION,
        "href": "admin-imports.html?deliveryPackageStatus=ready",
    },
]
IMPORT_BATCH_FINE_SCOPE_FIELDS = (
    ("towns", "townCode"),
    ("villages", "villageCode"),
    ("blockCodes", "blockCode"),
)
POSTGIS_IMPORT_BATCH_COLUMNS = [
    "id",
    "file_name",
    "file_type",
    "status",
    "total_rows",
    "valid_rows",
    "invalid_rows",
    "created_by",
    "report_json",
    "created_at",
    "completed_at",
]
ROOT_DIR = Path(__file__).resolve().parents[2]
IMPORT_SOURCE_ROOTS = [
    ROOT_DIR / "converted-data",
    ROOT_DIR / "data" / "samples",
    ROOT_DIR / "data" / "remote-sensing" / "inbox",
]
SUPPORTED_IMPORT_SOURCE_SUFFIXES = {
    ".geojson",
    ".json",
    ".csv",
    ".xlsx",
    ".zip",
    ".kml",
    ".kmz",
    ".ovkml",
    ".ovkmz",
    ".ovobj",
}


def require_map_layer_upsert_permissions(
    context: AuthContext,
    *,
    existing: dict[str, Any],
    next_layer: dict[str, Any],
) -> None:
    if existing:
        require_permission(context, MAP_LAYER_PERMISSIONS["update"])
    else:
        require_permission(context, MAP_LAYER_PERMISSIONS["create"])
    if bool(next_layer.get("visibleOnDashboard")) or str(next_layer.get("status") or "") == "published":
        require_permission(context, MAP_LAYER_PERMISSIONS["publish"])


def existing_import_batch_layer(batch_id: str) -> dict[str, Any]:
    layer_id = f"import-layer-{batch_id}"
    record_code = f"IMPORT-LAYER-{batch_id}"
    for record in layer_records():
        if str(record.get("id")) == layer_id or str(record.get("recordCode")) == record_code:
            return record
    return {}


def require_import_batch_layer_preflight_permissions(context: AuthContext, batch_id: str) -> None:
    require_map_layer_upsert_permissions(
        context,
        existing=existing_import_batch_layer(batch_id),
        next_layer={"status": "published", "visibleOnDashboard": True},
    )


SQUARE_METERS_PER_MU = 666.6666667
NUMBER_PATTERN = re.compile(r"-?\d+(?:,\d{3})*(?:\.\d+)?|-?\d+(?:\.\d+)?")
TD_PAIR_PATTERN = re.compile(r"<td[^>]*>\s*(.*?)\s*</td>\s*<td[^>]*>\s*(.*?)\s*</td>", re.I | re.S)


class ImportSourceRequest(BaseModel):
    path: str
    strategy: str = "upsert"


class ImportBatchFilters(BaseModel):
    q: str = ""
    status: str = ""
    reviewStatus: str = ""
    acceptanceStatus: str = ""
    qualityStatus: str = ""
    publishRiskStatus: str = ""
    sceneId: str = ""
    workflowQueue: str = ""
    includeDeleted: bool = False
    limit: int = Field(default=100, ge=1, le=1000)
    offset: int = Field(default=0, ge=0)


class ImportBatchReviewRequest(BaseModel):
    decision: str = Field(min_length=1)
    comment: str = ""


class ImportBatchAcceptanceRequest(BaseModel):
    status: str = Field(min_length=1)
    comment: str = ""


class ImportQualityIssueUpdateRequest(BaseModel):
    status: str = Field(min_length=1)
    comment: str = ""


class ImportBatchSceneLayerRequest(BaseModel):
    sceneId: str = Field(min_length=1)
    relationType: str = "coverage"
    capturedAt: str | None = None
    confidence: float | None = None
    publishLayer: bool = True
    layerName: str | None = None
    zIndex: int | None = None
    style: dict[str, Any] | None = None
    properties: dict[str, Any] | None = None


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

FIELD_ALIASES: dict[str, list[str]] = {
    "blockCode": ["blockCode", "林班编号", "小班编号", "编号", "block_code"],
    "name": ["name", "林班名称", "小班名称", "名称"],
    "countyCode": ["countyCode", "区县编码"],
    "countyName": ["countyName", "区县", "县", "行政区"],
    "townCode": ["townCode", "乡镇编码"],
    "townName": ["townName", "乡镇", "镇"],
    "villageCode": ["villageCode", "村编码"],
    "villageName": ["villageName", "村", "行政村"],
    "areaMu": ["areaMu", "面积", "亩", "面积亩"],
    "qualityGrade": ["qualityGrade", "质量等级"],
    "healthStatus": ["healthStatus", "健康状态"],
    "riskLevel": ["riskLevel", "风险等级"],
    "baseType": ["baseType"],
    "operationType": ["operationType"],
    "forestType": ["forestType"],
    "ownershipStatus": ["ownershipStatus", "权属状态"],
    "managementStatus": ["managementStatus", "经营状态"],
}


def pick_field(values: dict[str, Any], field_name: str) -> Any:
    for alias in FIELD_ALIASES.get(field_name, []):
        if alias not in values:
            continue
        value = values[alias]
        if value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
        if value == "":
            continue
        return value
    return None


def normalize_geometry(geometry: dict[str, Any] | None) -> dict[str, Any] | None:
    if not geometry:
        return None
    geometry_type = geometry.get("type")
    coordinates = geometry.get("coordinates")
    if geometry_type == "Polygon":
        return {"type": "MultiPolygon", "coordinates": [coordinates]}
    if geometry_type == "MultiPolygon":
        return {"type": "MultiPolygon", "coordinates": coordinates}
    return None


def normalize_import_record(
    properties: dict[str, Any],
    geometry: dict[str, Any] | None = None,
) -> dict[str, Any]:
    errors: list[str] = []
    record: dict[str, Any] = {
        field_name: pick_field(properties, field_name) for field_name in FIELD_ALIASES
    }
    area_mu = record.get("areaMu")
    if area_mu is not None:
        try:
            record["areaMu"] = float(area_mu)
        except (TypeError, ValueError):
            errors.append("areaMu must be a number")
            record["areaMu"] = None
    record["geometry"] = normalize_geometry(geometry)
    record["properties"] = dict(properties)
    if errors:
        record[IMPORT_ERROR_FIELD] = errors
    return record


def clean_table_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = html.unescape(text)
    text = text.replace("\xa0", " ").strip()
    lines = [line.strip() for line in re.split(r"[\r\n]+", text) if line.strip()]
    cleaned = lines[-1] if lines else text
    if cleaned in {"<空>", "空", "NULL", "null", "None", "none"}:
        return ""
    return cleaned


def extract_html_table_fields(markup: str | None) -> dict[str, Any]:
    if not markup:
        return {}
    fields: dict[str, Any] = {}
    for raw_key, raw_value in TD_PAIR_PATTERN.findall(markup):
        key = clean_table_text(raw_key)
        value = clean_table_text(raw_value)
        if key:
            fields[key] = value
    return fields


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def child_text(element: ET.Element, name: str) -> str:
    for child in list(element):
        if local_name(child.tag) == name:
            return clean_table_text(child.text)
    return ""


def child_raw_text(element: ET.Element, name: str) -> str:
    for child in list(element):
        if local_name(child.tag) == name:
            return child.text or ""
    return ""


def descendant_text(element: ET.Element, name: str) -> str:
    for child in element.iter():
        if child is not element and local_name(child.tag) == name:
            return clean_table_text(child.text)
    return ""


def field_value(fields: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        value = fields.get(key)
        if value not in (None, ""):
            return value
    folded = {str(key).casefold(): value for key, value in fields.items()}
    for key in keys:
        value = folded.get(key.casefold())
        if value not in (None, ""):
            return value
    return None


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    match = NUMBER_PATTERN.search(str(value).replace(",", ""))
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def normalize_area_mu(value: Any, *, field_name: str = "") -> float | None:
    number = parse_number(value)
    if number is None:
        return None
    text = str(value)
    normalized_field = field_name.casefold()
    is_square_meters = (
        normalized_field in {"xbmj", "xbtmmj"}
        or "\u5e73\u65b9\u7c73" in text
        or "m2" in text.casefold()
        or "\u33a1" in text
    )
    if is_square_meters:
        return number / SQUARE_METERS_PER_MU
    return number


def area_mu_from_fields(fields: dict[str, Any]) -> float | None:
    for key in ["areaMu", "FZMJ", "XBMJ", "XBTMMJ", "\u9762\u79ef"]:
        value = field_value(fields, [key])
        area_mu = normalize_area_mu(value, field_name=key)
        if area_mu is not None:
            return area_mu
    return None


def source_file_type(file_name: str) -> str:
    return Path(file_name).suffix.lower().lstrip(".") or "upload"


def build_rights_payload(fields: dict[str, Any]) -> dict[str, Any]:
    holder = field_value(
        fields,
        [
            "holder",
            "FBF",
            "LMSHIYQRMC",
            "\u6743\u5229\u4eba",
            "\u53d1\u5305\u65b9",
            "\u7ecf\u8425\u4e3b\u4f53",
        ],
    )
    certificate_no = field_value(
        fields,
        [
            "certificateNo",
            "YZDBH",
            "\u6797\u6743\u8bc1\u53f7",
            "\u8bc1\u4e66\u7f16\u53f7",
        ],
    )
    right_end = field_value(fields, ["SYQJSSJ", "\u6743\u5229\u7ed3\u675f\u65f6\u95f4"])
    right_type = field_value(fields, ["QLXZ", "\u6743\u5229\u6027\u8d28"])
    rights = {
        "holder": holder,
        "certificateNo": certificate_no,
        "rightEnd": right_end,
        "rightType": right_type,
    }
    return {key: value for key, value in rights.items() if value not in (None, "")}


def build_external_record(
    *,
    fields: dict[str, Any],
    geometry: dict[str, Any] | None,
    file_name: str,
    source_format: str,
    placemark_name: str = "",
    row_number: int = 1,
) -> dict[str, Any]:
    stem = Path(file_name).stem
    block_code = field_value(
        fields,
        [
            "blockCode",
            "XBNO",
            "YZDBH",
            "\u540d\u79f0",
            "name",
        ],
    )
    block_code = clean_table_text(block_code or placemark_name or f"{stem}-{row_number:04d}")
    name = clean_table_text(placemark_name or field_value(fields, ["name", "\u540d\u79f0"]) or block_code)
    rights = build_rights_payload(fields)

    enriched = dict(fields)
    enriched.update(
        {
            "blockCode": block_code,
            "name": name,
            "countyCode": field_value(fields, ["countyCode", "XSQ", "JYXSQ"]),
            "townCode": field_value(fields, ["townCode", "XZC", "JYXZC"]),
            "villageCode": field_value(fields, ["villageCode", "CGQ", "JYCGQ"]),
            "areaMu": area_mu_from_fields(fields),
            "ownershipStatus": "certified" if rights.get("certificateNo") else "pending",
            "managementStatus": field_value(fields, ["managementStatus"]) or "active",
            "rights": {
                **rights,
                "archiveStatus": "complete" if rights.get("certificateNo") else "partial",
            },
            "source": {
                "fileName": file_name,
                "fileType": source_format,
                "placemarkName": placemark_name,
                "rowNumber": row_number,
                "geometryStatus": "available" if geometry else "not_found",
            },
            "rawAttributes": dict(fields),
        }
    )
    if enriched["areaMu"] is None:
        enriched.pop("areaMu")
    return normalize_import_record(enriched, geometry)


def parse_kml_coordinates(value: str | None) -> list[list[float]]:
    if not value:
        return []
    points: list[list[float]] = []
    for token in value.split():
        parts = token.split(",")
        if len(parts) < 2:
            continue
        try:
            lon = float(parts[0])
            lat = float(parts[1])
        except ValueError:
            continue
        if -180 <= lon <= 180 and -90 <= lat <= 90:
            points.append([lon, lat])
    if points and points[0] != points[-1]:
        points.append(points[0])
    return points if len(points) >= 4 else []


def parse_kml_polygon(polygon: ET.Element) -> list[list[list[float]]]:
    rings: list[list[list[float]]] = []
    for child in polygon.iter():
        if local_name(child.tag) == "coordinates":
            ring = parse_kml_coordinates(child.text)
            if ring:
                rings.append(ring)
    return rings


def parse_kml_geometry(placemark: ET.Element) -> dict[str, Any] | None:
    polygons: list[list[list[list[float]]]] = []
    for child in placemark.iter():
        if local_name(child.tag) != "Polygon":
            continue
        rings = parse_kml_polygon(child)
        if rings:
            polygons.append(rings)
    if not polygons:
        return None
    return {"type": "MultiPolygon", "coordinates": polygons}


def extract_extended_data_fields(placemark: ET.Element) -> dict[str, Any]:
    fields: dict[str, Any] = {}
    for child in placemark.iter():
        name = local_name(child.tag)
        if name == "Data":
            key = child.attrib.get("name")
            value = descendant_text(child, "value")
            if key and value:
                fields[clean_table_text(key)] = value
        elif name == "SimpleData":
            key = child.attrib.get("name")
            value = clean_table_text(child.text)
            if key and value:
                fields[clean_table_text(key)] = value
    return fields


def parse_kml_file(file_name: str, content: bytes, *, source_format: str | None = None) -> list[dict[str, Any]]:
    try:
        root = ET.fromstring(content.decode("utf-8-sig", errors="replace"))
    except ET.ParseError as exc:
        raise HTTPException(status_code=400, detail="KML import file is not valid XML") from exc

    records: list[dict[str, Any]] = []
    seen_codes: dict[str, int] = {}
    file_type = source_format or source_file_type(file_name)
    for row_number, placemark in enumerate(
        [element for element in root.iter() if local_name(element.tag) == "Placemark"],
        start=1,
    ):
        placemark_name = child_text(placemark, "name")
        fields = extract_html_table_fields(child_raw_text(placemark, "description"))
        fields.update(extract_extended_data_fields(placemark))
        geometry = parse_kml_geometry(placemark)
        if not fields and not placemark_name and not geometry:
            continue
        record = build_external_record(
            fields=fields,
            geometry=geometry,
            file_name=file_name,
            source_format=file_type,
            placemark_name=placemark_name,
            row_number=row_number,
        )
        code = record.get("blockCode")
        if code:
            seen_codes[code] = seen_codes.get(code, 0) + 1
            if seen_codes[code] > 1:
                record["blockCode"] = f"{code}-{row_number}"
                record["properties"]["blockCode"] = record["blockCode"]
        records.append(record)
    return records


def parse_kmz_file(file_name: str, content: bytes) -> list[dict[str, Any]]:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            kml_names = [name for name in archive.namelist() if name.lower().endswith((".kml", ".ovkml"))]
            if not kml_names:
                raise HTTPException(status_code=400, detail="KMZ import must include a KML file")
            preferred = next((name for name in kml_names if Path(name).name.lower() == "doc.kml"), kml_names[0])
            return parse_kml_file(file_name, archive.read(preferred), source_format=source_file_type(file_name))
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=400, detail="KMZ import file is not a valid ZIP archive") from exc


def parse_ovobj_file(file_name: str, content: bytes) -> list[dict[str, Any]]:
    if not content.startswith(b"OviO"):
        raise HTTPException(status_code=400, detail="OVOBJ import file is not an Ovi object file")
    text = content.decode("utf-8-sig", errors="ignore")
    fields = extract_html_table_fields(text)
    if not fields:
        raise HTTPException(status_code=400, detail="OVOBJ import did not contain a readable attribute table")
    return [
        build_external_record(
            fields=fields,
            geometry=None,
            file_name=file_name,
            source_format="ovobj",
            placemark_name=Path(file_name).stem,
            row_number=1,
        )
    ]


def parse_geojson(content: bytes) -> list[dict[str, Any]]:
    payload = json.loads(content.decode("utf-8-sig"))
    if isinstance(payload, dict) and payload.get("type") == "FeatureCollection":
        features = payload.get("features") or []
    elif isinstance(payload, dict) and payload.get("type") == "Feature":
        features = [payload]
    elif isinstance(payload, list):
        features = payload
    else:
        raise HTTPException(status_code=400, detail="GeoJSON import must be a FeatureCollection, Feature, or JSON list")

    return [
        normalize_import_record(feature.get("properties") or {}, feature.get("geometry"))
        for feature in features
    ]


def parse_csv_file(content: bytes) -> list[dict[str, Any]]:
    rows = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    return [normalize_import_record(dict(row)) for row in rows]


def parse_xlsx_file(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=503, detail="XLSX import requires openpyxl") from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return []
    normalized_headers = [str(value).strip() if value is not None else "" for value in headers]
    records: list[dict[str, Any]] = []
    for row in rows:
        values = {
            normalized_headers[index]: row[index]
            for index in range(min(len(normalized_headers), len(row)))
            if normalized_headers[index]
        }
        records.append(normalize_import_record(values))
    return records


def parse_shapefile_zip(content: bytes) -> list[dict[str, Any]]:
    try:
        import pyogrio
        from shapely.geometry import mapping
    except ImportError as exc:
        raise HTTPException(status_code=503, detail="Shapefile import requires pyogrio and shapely") from exc

    with tempfile.TemporaryDirectory() as tmp_dir:
        root = Path(tmp_dir)
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            safe_extract_zip(archive, root)
        shapefiles = list(root.rglob("*.shp"))
        if not shapefiles:
            raise HTTPException(status_code=400, detail="Shapefile ZIP must include a .shp file")
        frame = pyogrio.read_dataframe(shapefiles[0])
        if getattr(frame, "crs", None):
            frame = frame.to_crs("EPSG:4326")

        records: list[dict[str, Any]] = []
        for _, row in frame.iterrows():
            values = row.to_dict()
            geometry = values.pop("geometry", None)
            records.append(
                normalize_import_record(
                    values,
                    mapping(geometry) if geometry is not None else None,
                )
            )
        return records


def safe_extract_zip(archive: zipfile.ZipFile, root: Path) -> None:
    root = root.resolve()
    for member in archive.infolist():
        member_path = Path(member.filename)
        if member_path.is_absolute() or ".." in member_path.parts:
            raise HTTPException(status_code=400, detail="Shapefile ZIP contains an unsafe path")
        target = (root / member.filename).resolve()
        if root != target and root not in target.parents:
            raise HTTPException(status_code=400, detail="Shapefile ZIP contains an unsafe path")
        archive.extract(member, root)


def parse_import_file(file_name: str, content: bytes) -> list[dict[str, Any]]:
    lower_name = file_name.lower()
    if lower_name.endswith(".geojson") or lower_name.endswith(".json"):
        return parse_geojson(content)
    if lower_name.endswith(".csv"):
        return parse_csv_file(content)
    if lower_name.endswith(".xlsx"):
        return parse_xlsx_file(content)
    if lower_name.endswith((".kml", ".ovkml")):
        return parse_kml_file(file_name, content)
    if lower_name.endswith((".kmz", ".ovkmz")):
        return parse_kmz_file(file_name, content)
    if lower_name.endswith(".ovobj"):
        return parse_ovobj_file(file_name, content)
    if lower_name.endswith(".zip"):
        return parse_shapefile_zip(content)
    raise HTTPException(
        status_code=400,
        detail="Supported formats: CSV, XLSX, GeoJSON, JSON, KML, KMZ, OVKML, OVKMZ, OVOBJ, Shapefile ZIP",
    )


def path_is_relative_to(path: Path, root: Path) -> bool:
    return path == root or root in path.parents


def allowed_import_source_roots() -> list[Path]:
    return [root.expanduser().resolve() for root in IMPORT_SOURCE_ROOTS if root.expanduser().exists()]


def display_source_path(path: Path) -> str:
    resolved = path.resolve()
    root = ROOT_DIR.resolve()
    if path_is_relative_to(resolved, root):
        return resolved.relative_to(root).as_posix()
    return str(resolved)


def import_source_record(path: Path, root: Path) -> dict[str, Any]:
    stat = path.stat()
    suffix = path.suffix.lower().lstrip(".")
    return {
        "path": display_source_path(path),
        "fileName": path.name,
        "name": path.stem,
        "fileType": suffix,
        "sizeBytes": stat.st_size,
        "updatedAt": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
        "sourceRoot": display_source_path(root),
    }


def list_import_source_files() -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for root in allowed_import_source_roots():
        for path in root.rglob("*"):
            if not path.is_file() or path.suffix.lower() not in SUPPORTED_IMPORT_SOURCE_SUFFIXES:
                continue
            items.append(import_source_record(path.resolve(), root))
    return sorted(items, key=lambda item: (item["updatedAt"], item["path"]), reverse=True)


def resolve_import_source_path(value: str) -> Path:
    raw = Path(str(value or "").strip())
    if not str(raw):
        raise HTTPException(status_code=400, detail="path is required")

    candidates = [raw.expanduser().resolve()] if raw.is_absolute() else [(ROOT_DIR / raw).resolve()]
    roots = allowed_import_source_roots()
    for candidate in candidates:
        if candidate.suffix.lower() not in SUPPORTED_IMPORT_SOURCE_SUFFIXES:
            continue
        if not candidate.is_file():
            continue
        if any(path_is_relative_to(candidate, root) for root in roots):
            return candidate
    raise HTTPException(status_code=400, detail="path is not an allowed import source")


def validate_record(record: dict[str, Any]) -> list[str]:
    errors: list[str] = list(record.get(IMPORT_ERROR_FIELD) or [])
    if not record.get("blockCode"):
        errors.append("blockCode is required")
    if not record.get("name"):
        errors.append("name is required")
    return errors


def clean_import_record(record: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in record.items() if key != IMPORT_ERROR_FIELD}


def build_report(
    *,
    batch_id: str,
    file_name: str,
    total_rows: int,
    valid_rows: int,
    invalid_rows: int,
    errors: list[dict[str, Any]],
    imported_blocks: list[dict[str, Any]] | None = None,
    imported_rights_archives: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    suffix = Path(file_name).suffix.lower().lstrip(".")
    timestamp = now_iso()
    return {
        "id": batch_id,
        "fileName": file_name,
        "fileType": suffix,
        "status": "completed",
        "totalRows": total_rows,
        "validRows": valid_rows,
        "invalidRows": invalid_rows,
        "errors": errors,
        "importedBlocks": imported_blocks or [],
        "importedRightsArchives": imported_rights_archives or [],
        "imageryLinks": [],
        "rolledBackBlocks": [],
        "rollbackSummary": {},
        "reviewStatus": "pending",
        "reviewComment": "",
        "reviewedAt": None,
        "reviewedBy": "",
        "reviewEvents": [],
        "acceptanceStatus": "pending",
        "acceptanceComment": "",
        "acceptedAt": None,
        "acceptedBy": "",
        "acceptanceEvents": [],
        "auditEvents": [],
        "qualityStatus": "pending",
        "qualityFindings": [],
        "reviewRecommendation": "",
        "publishRiskStatus": "unknown",
        "createdAt": timestamp,
        "completedAt": timestamp,
    }


def json_value(value: Any, default: Any) -> Any:
    if value is None:
        return default
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return default
    return value


def datetime_to_iso(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def mysql_datetime(value: Any) -> Any:
    if not isinstance(value, str) or not value.strip():
        return value or None
    try:
        parsed = datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
    except ValueError:
        return value
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed


def postgis_connect():
    try:
        import psycopg
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"import batch PostGIS database requires psycopg. {exc}") from exc

    try:
        return psycopg.connect(get_settings().database_url)
    except Exception as exc:
        raise HTTPException(status_code=503, detail="import batch PostGIS database is unavailable") from exc


def normalize_postgis_import_batch_row(row: Any) -> dict[str, Any]:
    if hasattr(row, "keys"):
        source = dict(row)
    else:
        source = dict(zip(POSTGIS_IMPORT_BATCH_COLUMNS, row))
    report = json_value(source.get("report_json"), {})
    if not isinstance(report, dict):
        report = {}
    report.setdefault("id", str(source.get("id") or ""))
    report.setdefault("fileName", source.get("file_name") or "")
    report.setdefault("fileType", source.get("file_type") or "")
    report.setdefault("status", source.get("status") or "completed")
    report.setdefault("totalRows", source.get("total_rows") or 0)
    report.setdefault("validRows", source.get("valid_rows") or 0)
    report.setdefault("invalidRows", source.get("invalid_rows") or 0)
    report.setdefault("errors", [])
    report.setdefault("importedBlocks", [])
    report.setdefault("importedRightsArchives", [])
    report.setdefault("imageryLinks", [])
    report.setdefault("rolledBackBlocks", [])
    report.setdefault("rollbackSkippedBlocks", [])
    report.setdefault("rollbackSummary", {})
    report.setdefault("createdBy", source.get("created_by"))
    report.setdefault("createdAt", datetime_to_iso(source.get("created_at")))
    report.setdefault("completedAt", datetime_to_iso(source.get("completed_at")))
    report.setdefault("reviewStatus", "pending")
    report.setdefault("reviewComment", "")
    report.setdefault("reviewedAt", None)
    report.setdefault("reviewedBy", "")
    report.setdefault("reviewEvents", [])
    report.setdefault("acceptanceStatus", "pending")
    report.setdefault("acceptanceComment", "")
    report.setdefault("acceptedAt", None)
    report.setdefault("acceptedBy", "")
    report.setdefault("acceptanceEvents", [])
    report["acceptanceEvents"] = json_value(report.get("acceptanceEvents"), [])
    if not isinstance(report["acceptanceEvents"], list):
        report["acceptanceEvents"] = []
    report.setdefault("auditEvents", [])
    report.setdefault("qualityIssueEvents", [])
    report.setdefault("qualityStatus", "pending")
    report.setdefault("qualityFindings", [])
    report.setdefault("reviewRecommendation", "")
    report.setdefault("publishRiskStatus", "unknown")
    report.setdefault("rolledBackAt", None)
    report.setdefault("deletedAt", None)
    return report


def load_import_report_postgis(batch_id: str) -> dict[str, Any] | None:
    with postgis_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    id::text,
                    file_name,
                    file_type,
                    status,
                    total_rows,
                    valid_rows,
                    invalid_rows,
                    created_by,
                    report_json,
                    created_at,
                    completed_at
                FROM import_batches
                WHERE id = %s
                """,
                (batch_id,),
            )
            row = cur.fetchone()
    return normalize_postgis_import_batch_row(row) if row else None


def postgis_import_filter_sql(filters: ImportBatchFilters) -> tuple[str, list[Any]]:
    clauses = []
    params: list[Any] = []
    has_scene_link_sql = """
        EXISTS (
            SELECT 1
            FROM jsonb_array_elements(COALESCE(report_json->'imageryLinks', '[]'::jsonb)) AS link
            WHERE COALESCE(link->>'sceneId', '') <> ''
        )
    """
    if not filters.includeDeleted:
        clauses.append("COALESCE(report_json->>'deletedAt', '') = ''")
    if filters.status:
        clauses.append("status = %s")
        params.append(filters.status)
    if filters.reviewStatus:
        clauses.append("COALESCE(report_json->>'reviewStatus', 'pending') = %s")
        params.append(filters.reviewStatus)
    if filters.acceptanceStatus:
        clauses.append("COALESCE(report_json->>'acceptanceStatus', 'pending') = %s")
        params.append(filters.acceptanceStatus)
    if filters.qualityStatus:
        clauses.append("COALESCE(report_json->>'qualityStatus', 'pending') = %s")
        params.append(filters.qualityStatus)
    if filters.publishRiskStatus:
        clauses.append("COALESCE(report_json->>'publishRiskStatus', 'unknown') = %s")
        params.append(filters.publishRiskStatus)
    if filters.q:
        query_text = f"%{filters.q}%"
        clauses.append("(file_name ILIKE %s OR file_type ILIKE %s OR report_json::text ILIKE %s)")
        params.extend([query_text, query_text, query_text])
    if filters.sceneId:
        clauses.append("COALESCE(report_json->'imageryLinks', '[]'::jsonb) @> %s::jsonb")
        params.append(json.dumps([{"sceneId": filters.sceneId}], ensure_ascii=False))
    workflow_queue = filters.workflowQueue.strip()
    if workflow_queue == "pendingReview":
        clauses.append("COALESCE(report_json->>'reviewStatus', 'pending') = 'pending'")
    elif workflow_queue == "approvedUnlinked":
        clauses.append(f"COALESCE(report_json->>'reviewStatus', 'pending') = 'approved' AND NOT {has_scene_link_sql}")
    elif workflow_queue == "readyForLayerLink":
        clauses.append(
            f"""
            COALESCE(report_json->>'reviewStatus', 'pending') = 'approved'
            AND NOT {has_scene_link_sql}
            AND valid_rows > 0
            AND COALESCE(report_json->>'qualityStatus', 'pending') IN ('passed', 'warning')
            AND COALESCE(report_json->>'publishRiskStatus', 'unknown') IN ('clear', 'ready', 'warning')
            """
        )
    elif workflow_queue == "linked":
        clauses.append(has_scene_link_sql)
    elif workflow_queue:
        clauses.append("FALSE")
    return (" WHERE " + " AND ".join(clauses)) if clauses else "", params


def list_import_reports_postgis(filters: ImportBatchFilters, *, paginate: bool = True) -> list[dict[str, Any]]:
    where_sql, params = postgis_import_filter_sql(filters)
    pagination_sql = "LIMIT %s OFFSET %s" if paginate else ""
    sql = f"""
        SELECT
            id::text,
            file_name,
            file_type,
            status,
            total_rows,
            valid_rows,
            invalid_rows,
            created_by,
            report_json,
            created_at,
            completed_at
        FROM import_batches
        {where_sql}
        ORDER BY created_at DESC, file_name
        {pagination_sql}
    """
    if paginate:
        params.extend([filters.limit, filters.offset])
    with postgis_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, tuple(params))
            return [normalize_postgis_import_batch_row(row) for row in cur.fetchall()]


def count_import_reports_postgis(filters: ImportBatchFilters) -> int:
    where_sql, params = postgis_import_filter_sql(filters)
    with postgis_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) FROM import_batches{where_sql}", tuple(params))
            row = cur.fetchone()
    return int(row[0] if row else 0)


def upsert_import_report_postgis(report: dict[str, Any]) -> None:
    with postgis_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO import_batches (
                    id,
                    file_name,
                    file_type,
                    status,
                    total_rows,
                    valid_rows,
                    invalid_rows,
                    created_by,
                    report_json,
                    created_at,
                    completed_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s, %s)
                ON CONFLICT (id) DO UPDATE SET
                    file_name = EXCLUDED.file_name,
                    file_type = EXCLUDED.file_type,
                    status = EXCLUDED.status,
                    total_rows = EXCLUDED.total_rows,
                    valid_rows = EXCLUDED.valid_rows,
                    invalid_rows = EXCLUDED.invalid_rows,
                    created_by = EXCLUDED.created_by,
                    report_json = EXCLUDED.report_json,
                    completed_at = EXCLUDED.completed_at
                """,
                (
                    report.get("id"),
                    report.get("fileName"),
                    report.get("fileType"),
                    report.get("status"),
                    report.get("totalRows"),
                    report.get("validRows"),
                    report.get("invalidRows"),
                    report.get("createdBy"),
                    json.dumps(report, ensure_ascii=False),
                    report.get("createdAt"),
                    report.get("completedAt"),
                ),
            )
        conn.commit()


MYSQL_IMPORT_SELECT_SQL = """
    SELECT
        id, file_name, file_type, status,
        total_rows, valid_rows, invalid_rows,
        created_by, report_json, created_at, completed_at
    FROM import_batches
"""

MYSQL_IMPORT_LIST_SELECT_SQL = """
    SELECT
        id, file_name, file_type, status,
        total_rows, valid_rows, invalid_rows,
        created_by,
        JSON_REMOVE(report_json, '$.importedBlocks', '$.importedRightsArchives') AS report_json,
        created_at, completed_at
    FROM import_batches
"""


def normalize_mysql_import_block_target(row: Any) -> dict[str, Any]:
    if hasattr(row, "keys"):
        source = dict(row)
        raw_target = source.get("target_json")
        block_code = source.get("block_code")
        name = source.get("name")
        action = source.get("import_action")
    else:
        raw_target, block_code, name, action = row
    target = json_value(raw_target, {})
    if not isinstance(target, dict):
        target = {}
    target.setdefault("blockCode", block_code or "")
    target.setdefault("name", name or "")
    target.setdefault("action", action or "")
    return target


def normalize_mysql_import_right_target(row: Any) -> dict[str, Any]:
    if hasattr(row, "keys"):
        source = dict(row)
        raw_target = source.get("target_json")
        archive_code = source.get("archive_code")
    else:
        raw_target, archive_code = row
    target = json_value(raw_target, {})
    if not isinstance(target, dict):
        target = {}
    target.setdefault("archiveCode", archive_code or "")
    return target


def load_import_report_mysql(
    batch_id: str,
    *,
    include_targets: bool = True,
) -> dict[str, Any] | None:
    with mysql_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(f"{MYSQL_IMPORT_SELECT_SQL} WHERE id = %s", (batch_id,))
            row = cur.fetchone()
            if not row:
                return None
            report = normalize_postgis_import_batch_row(row)
            if not include_targets:
                report["_targetsLoaded"] = False
                return report
            cur.execute(
                "SELECT links.target_json, blocks.block_code, blocks.name, links.import_action "
                "FROM import_batch_block_links links "
                "JOIN forest_blocks blocks ON blocks.id = links.forest_block_id "
                "WHERE links.import_batch_id = %s ORDER BY blocks.block_code",
                (batch_id,),
            )
            report["importedBlocks"] = [
                normalize_mysql_import_block_target(target_row)
                for target_row in cur.fetchall()
            ]
            cur.execute(
                "SELECT links.target_json, rights.archive_code "
                "FROM import_batch_right_links links "
                "JOIN forest_rights rights ON rights.id = links.forest_right_id "
                "WHERE links.import_batch_id = %s ORDER BY rights.archive_code",
                (batch_id,),
            )
            report["importedRightsArchives"] = [
                normalize_mysql_import_right_target(target_row)
                for target_row in cur.fetchall()
            ]
    return report


def mysql_import_scope_exists_sql(
    context: AuthContext | None,
    params: list[Any],
) -> str:
    if context is None or not context_has_import_batch_scope(context):
        return ""
    scope_clauses = ["scope_links.import_batch_id = import_batches.id"]
    if has_effective_area_scope(context):
        area_values = sorted(value for value in effective_areas(context) if value and value != "*")
        if not area_values:
            scope_clauses.append("FALSE")
        else:
            scope_clauses.append(
                f"scope_blocks.county_code IN ({', '.join(['%s'] * len(area_values))})"
            )
            params.extend(area_values)
    scope_columns = {
        "towns": "town_code",
        "villages": "village_code",
        "blockCodes": "block_code",
    }
    for scope_key, column_name in scope_columns.items():
        values = sorted(
            value
            for value in effective_data_scope_values(context, scope_key)
            if value and value != "*"
        )
        if values:
            scope_clauses.append(
                f"scope_blocks.{column_name} IN ({', '.join(['%s'] * len(values))})"
            )
            params.extend(values)
    return (
        "EXISTS (SELECT 1 FROM import_batch_block_links scope_links "
        "JOIN forest_blocks scope_blocks ON scope_blocks.id = scope_links.forest_block_id "
        f"WHERE {' AND '.join(scope_clauses)})"
    )


def mysql_import_filter_sql(
    filters: ImportBatchFilters,
    context: AuthContext | None = None,
) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    params: list[Any] = []
    has_scene_link_sql = "EXISTS (SELECT 1 FROM import_batch_scene_links isl WHERE isl.import_batch_id = import_batches.id)"
    if not filters.includeDeleted:
        clauses.append("deleted_at IS NULL")
    if filters.status:
        clauses.append("status = %s")
        params.append(filters.status)
    if filters.reviewStatus:
        clauses.append("review_status = %s")
        params.append(filters.reviewStatus)
    if filters.acceptanceStatus:
        clauses.append("acceptance_status = %s")
        params.append(filters.acceptanceStatus)
    if filters.qualityStatus:
        clauses.append("quality_status = %s")
        params.append(filters.qualityStatus)
    if filters.publishRiskStatus:
        clauses.append("publish_risk_status = %s")
        params.append(filters.publishRiskStatus)
    if filters.q:
        query_text = f"%{filters.q}%"
        clauses.append(
            "(file_name LIKE %s OR file_type LIKE %s OR CAST(report_json AS CHAR) LIKE %s "
            "OR EXISTS (SELECT 1 FROM import_batch_block_links search_block_links "
            "JOIN forest_blocks search_blocks ON search_blocks.id = search_block_links.forest_block_id "
            "WHERE search_block_links.import_batch_id = import_batches.id "
            "AND (search_blocks.block_code LIKE %s OR search_blocks.name LIKE %s)) "
            "OR EXISTS (SELECT 1 FROM import_batch_right_links search_right_links "
            "JOIN forest_rights search_rights ON search_rights.id = search_right_links.forest_right_id "
            "WHERE search_right_links.import_batch_id = import_batches.id "
            "AND (search_rights.archive_code LIKE %s OR search_rights.certificate_no LIKE %s "
            "OR search_rights.holder LIKE %s)))"
        )
        params.extend([query_text] * 8)
    if filters.sceneId:
        clauses.append(
            "EXISTS (SELECT 1 FROM import_batch_scene_links isl "
            "WHERE isl.import_batch_id = import_batches.id AND isl.scene_id = %s)"
        )
        params.append(filters.sceneId)
    workflow_queue = filters.workflowQueue.strip()
    if workflow_queue == "pendingReview":
        clauses.append("review_status = 'pending'")
    elif workflow_queue == "approvedUnlinked":
        clauses.append(f"review_status = 'approved' AND NOT {has_scene_link_sql}")
    elif workflow_queue == "readyForLayerLink":
        clauses.append(
            f"review_status = 'approved' AND NOT {has_scene_link_sql} AND valid_rows > 0 "
            "AND quality_status IN ('passed', 'warning') "
            "AND publish_risk_status IN ('clear', 'ready', 'warning')"
        )
    elif workflow_queue == "linked":
        clauses.append(has_scene_link_sql)
    elif workflow_queue:
        clauses.append("FALSE")
    scope_exists_sql = mysql_import_scope_exists_sql(context, params)
    if scope_exists_sql:
        clauses.append(scope_exists_sql)
    return (" WHERE " + " AND ".join(clauses)) if clauses else "", params


def list_import_reports_mysql(
    filters: ImportBatchFilters,
    *,
    paginate: bool = True,
    context: AuthContext | None = None,
) -> list[dict[str, Any]]:
    where_sql, params = mysql_import_filter_sql(filters, context)
    sql = f"{MYSQL_IMPORT_LIST_SELECT_SQL}{where_sql} ORDER BY created_at DESC, file_name"
    if paginate:
        sql += " LIMIT %s OFFSET %s"
        params.extend([filters.limit, filters.offset])
    with mysql_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, tuple(params))
            return [normalize_postgis_import_batch_row(row) for row in cur.fetchall()]


def hydrate_mysql_import_report_targets(
    reports: list[dict[str, Any]],
    *,
    batch_size: int = 500,
) -> list[dict[str, Any]]:
    reports_by_id = {
        str(report.get("id") or ""): report
        for report in reports
        if str(report.get("id") or "")
    }
    batch_ids = list(reports_by_id)
    for report in reports_by_id.values():
        report["importedBlocks"] = []
        report["importedRightsArchives"] = []
    if not batch_ids:
        return reports
    batch_size = max(1, int(batch_size))
    with mysql_connect() as conn:
        with conn.cursor() as cur:
            for start in range(0, len(batch_ids), batch_size):
                batch = batch_ids[start : start + batch_size]
                placeholders = ", ".join(["%s"] * len(batch))
                cur.execute(
                    "SELECT links.import_batch_id, links.target_json, blocks.block_code, "
                    "blocks.name, links.import_action "
                    "FROM import_batch_block_links links "
                    "JOIN forest_blocks blocks ON blocks.id = links.forest_block_id "
                    f"WHERE links.import_batch_id IN ({placeholders}) "
                    "ORDER BY links.import_batch_id, blocks.block_code",
                    tuple(batch),
                )
                for row in cur.fetchall():
                    if hasattr(row, "keys"):
                        source = dict(row)
                        batch_id = str(source.get("import_batch_id") or "")
                        target_row = (
                            source.get("target_json"),
                            source.get("block_code"),
                            source.get("name"),
                            source.get("import_action"),
                        )
                    else:
                        batch_id = str(row[0] or "")
                        target_row = row[1:]
                    report = reports_by_id.get(batch_id)
                    if report is not None:
                        report["importedBlocks"].append(
                            normalize_mysql_import_block_target(target_row)
                        )
                cur.execute(
                    "SELECT links.import_batch_id, links.target_json, rights.archive_code "
                    "FROM import_batch_right_links links "
                    "JOIN forest_rights rights ON rights.id = links.forest_right_id "
                    f"WHERE links.import_batch_id IN ({placeholders}) "
                    "ORDER BY links.import_batch_id, rights.archive_code",
                    tuple(batch),
                )
                for row in cur.fetchall():
                    if hasattr(row, "keys"):
                        source = dict(row)
                        batch_id = str(source.get("import_batch_id") or "")
                        target_row = (
                            source.get("target_json"),
                            source.get("archive_code"),
                        )
                    else:
                        batch_id = str(row[0] or "")
                        target_row = row[1:]
                    report = reports_by_id.get(batch_id)
                    if report is not None:
                        report["importedRightsArchives"].append(
                            normalize_mysql_import_right_target(target_row)
                        )
    return reports


def count_import_reports_mysql(
    filters: ImportBatchFilters,
    context: AuthContext | None = None,
) -> int:
    where_sql, params = mysql_import_filter_sql(filters, context)
    with mysql_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) FROM import_batches{where_sql}", tuple(params))
            row = cur.fetchone()
    return int(row[0] if row else 0)


def sync_import_batch_events_mysql(cur: Any, report: dict[str, Any]) -> None:
    batch_id = str(report.get("id") or "")
    cur.execute("DELETE FROM import_batch_events WHERE import_batch_id = %s", (batch_id,))
    event_collections = [
        ("audit", report.get("auditEvents") or []),
        ("review", report.get("reviewEvents") or []),
        ("acceptance", report.get("acceptanceEvents") or []),
        ("quality", report.get("qualityIssueEvents") or []),
    ]
    for event_type, events in event_collections:
        for index, raw_event in enumerate(events, start=1):
            if not isinstance(raw_event, dict):
                continue
            event = dict(raw_event)
            event_id = str(event.get("eventId") or event.get("id") or f"{event_type}-{index}")[:191]
            status = str(
                event.get("status")
                or (
                    report.get("reviewStatus")
                    if event_type == "review"
                    else report.get("acceptanceStatus")
                    if event_type == "acceptance"
                    else report.get("qualityStatus")
                    if event_type == "quality"
                    else report.get("status")
                )
                or ""
            )
            summary = str(event.get("summary") or event.get("message") or event.get("comment") or "")[:512]
            cur.execute(
                """
                INSERT INTO import_batch_events (
                    import_batch_id, event_type, event_id, action, status,
                    actor, event_at, summary, event_json
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    batch_id,
                    event_type,
                    event_id,
                    str(event.get("action") or ""),
                    status,
                    str(event.get("actor") or event.get("reviewedBy") or event.get("acceptedBy") or ""),
                    mysql_datetime(event.get("at") or event.get("createdAt") or event.get("updatedAt")),
                    summary,
                    json.dumps(event, ensure_ascii=False),
                ),
            )


def sync_import_batch_target_links_mysql(
    cur: Any,
    report: dict[str, Any],
    *,
    batch_size: int = 500,
) -> None:
    batch_id = str(report.get("id") or "")
    batch_size = max(1, int(batch_size))
    block_targets = [
        (
            str(item.get("blockCode") or "").strip(),
            str(item.get("action") or ""),
            json.dumps(item, ensure_ascii=False),
        )
        for item in report.get("importedBlocks") or []
        if isinstance(item, dict) and str(item.get("blockCode") or "").strip()
    ]
    right_targets = [
        (
            str(item.get("archiveCode") or "").strip(),
            json.dumps(item, ensure_ascii=False),
        )
        for item in report.get("importedRightsArchives") or []
        if isinstance(item, dict) and str(item.get("archiveCode") or "").strip()
    ]

    cur.execute("DELETE FROM import_batch_block_links WHERE import_batch_id = %s", (batch_id,))
    for start in range(0, len(block_targets), batch_size):
        batch = block_targets[start : start + batch_size]
        source_sql = " UNION ALL ".join(
            ["SELECT %s AS block_code, %s AS import_action, %s AS target_json"] * len(batch)
        )
        params: list[Any] = [batch_id]
        for block_code, action, target_json in batch:
            params.extend([block_code, action, target_json])
        cur.execute(
            "INSERT IGNORE INTO import_batch_block_links "
            "(import_batch_id, forest_block_id, import_action, target_json) "
            "SELECT %s, b.id, source.import_action, source.target_json "
            f"FROM ({source_sql}) source "
            "JOIN forest_blocks b ON b.block_code = source.block_code",
            tuple(params),
        )

    cur.execute("DELETE FROM import_batch_right_links WHERE import_batch_id = %s", (batch_id,))
    for start in range(0, len(right_targets), batch_size):
        batch = right_targets[start : start + batch_size]
        source_sql = " UNION ALL ".join(
            ["SELECT %s AS archive_code, %s AS target_json"] * len(batch)
        )
        params: list[Any] = [batch_id]
        for archive_code, target_json in batch:
            params.extend([archive_code, target_json])
        cur.execute(
            "INSERT IGNORE INTO import_batch_right_links "
            "(import_batch_id, forest_right_id, target_json) "
            "SELECT %s, r.id, source.target_json FROM "
            f"({source_sql}) source "
            "JOIN forest_rights r ON r.archive_code = source.archive_code",
            tuple(params),
        )


def compact_import_report_for_mysql(report: dict[str, Any]) -> dict[str, Any]:
    compact = {
        key: value
        for key, value in report.items()
        if not str(key).startswith("_")
    }
    targets_loaded = report.get("_targetsLoaded") is not False
    imported_blocks = compact.pop("importedBlocks", []) or []
    imported_rights = compact.pop("importedRightsArchives", []) or []
    if targets_loaded:
        compact["importedBlockCount"] = len(imported_blocks)
        compact["importedRightsArchiveCount"] = len(imported_rights)
    compact["targetsStorage"] = "relational"
    return compact


def public_import_report(report: dict[str, Any]) -> dict[str, Any]:
    return {
        key: value
        for key, value in report.items()
        if not str(key).startswith("_")
    }


def upsert_import_report_mysql(
    report: dict[str, Any],
    *,
    connection_factory: Any = mysql_connect,
) -> None:
    compact_report = compact_import_report_for_mysql(report)
    with connection_factory() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO import_batches (
                    id, file_name, file_type, status,
                    total_rows, valid_rows, invalid_rows,
                    review_status, acceptance_status, quality_status, publish_risk_status,
                    created_by, report_json, created_at, completed_at, deleted_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    file_name = VALUES(file_name),
                    file_type = VALUES(file_type),
                    status = VALUES(status),
                    total_rows = VALUES(total_rows),
                    valid_rows = VALUES(valid_rows),
                    invalid_rows = VALUES(invalid_rows),
                    review_status = VALUES(review_status),
                    acceptance_status = VALUES(acceptance_status),
                    quality_status = VALUES(quality_status),
                    publish_risk_status = VALUES(publish_risk_status),
                    created_by = VALUES(created_by),
                    report_json = VALUES(report_json),
                    completed_at = VALUES(completed_at),
                    deleted_at = VALUES(deleted_at)
                """,
                (
                    report.get("id"),
                    report.get("fileName"),
                    report.get("fileType"),
                    report.get("status"),
                    report.get("totalRows") or 0,
                    report.get("validRows") or 0,
                    report.get("invalidRows") or 0,
                    report.get("reviewStatus") or "pending",
                    report.get("acceptanceStatus") or "pending",
                    report.get("qualityStatus") or "pending",
                    report.get("publishRiskStatus") or "unknown",
                    report.get("createdBy"),
                    json.dumps(compact_report, ensure_ascii=False),
                    mysql_datetime(report.get("createdAt")),
                    mysql_datetime(report.get("completedAt")),
                    mysql_datetime(report.get("deletedAt")),
                ),
            )
            batch_id = str(report.get("id") or "")
            sync_import_batch_events_mysql(cur, report)
            if report.get("_targetsLoaded") is not False:
                sync_import_batch_target_links_mysql(cur, report)
            cur.execute("DELETE FROM import_batch_scene_links WHERE import_batch_id = %s", (batch_id,))
            for item in report.get("imageryLinks") or []:
                scene_id = str(item.get("sceneId") or "").strip() if isinstance(item, dict) else ""
                if scene_id:
                    cur.execute(
                        "INSERT IGNORE INTO import_batch_scene_links (import_batch_id, scene_id) VALUES (%s, %s)",
                        (batch_id, scene_id),
                    )
        conn.commit()


def normalize_import_report_json(report: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(report or {})
    normalized.setdefault("id", "")
    normalized.setdefault("fileName", "")
    normalized.setdefault("fileType", "")
    normalized.setdefault("status", "completed")
    normalized.setdefault("totalRows", 0)
    normalized.setdefault("validRows", 0)
    normalized.setdefault("invalidRows", 0)
    normalized["errors"] = json_value(normalized.get("errors"), [])
    if not isinstance(normalized["errors"], list):
        normalized["errors"] = []
    normalized["importedBlocks"] = json_value(normalized.get("importedBlocks"), [])
    if not isinstance(normalized["importedBlocks"], list):
        normalized["importedBlocks"] = []
    normalized["importedRightsArchives"] = json_value(normalized.get("importedRightsArchives"), [])
    if not isinstance(normalized["importedRightsArchives"], list):
        normalized["importedRightsArchives"] = []
    normalized["imageryLinks"] = json_value(normalized.get("imageryLinks"), [])
    if not isinstance(normalized["imageryLinks"], list):
        normalized["imageryLinks"] = []
    normalized["rolledBackBlocks"] = json_value(normalized.get("rolledBackBlocks"), [])
    if not isinstance(normalized["rolledBackBlocks"], list):
        normalized["rolledBackBlocks"] = []
    normalized["rollbackSummary"] = json_value(normalized.get("rollbackSummary"), {})
    if not isinstance(normalized["rollbackSummary"], dict):
        normalized["rollbackSummary"] = {}
    normalized.setdefault("reviewStatus", "pending")
    normalized.setdefault("reviewComment", "")
    normalized.setdefault("reviewedAt", None)
    normalized.setdefault("reviewedBy", "")
    normalized["reviewEvents"] = json_value(normalized.get("reviewEvents"), [])
    if not isinstance(normalized["reviewEvents"], list):
        normalized["reviewEvents"] = []
    normalized.setdefault("acceptanceStatus", "pending")
    normalized.setdefault("acceptanceComment", "")
    normalized.setdefault("acceptedAt", None)
    normalized.setdefault("acceptedBy", "")
    normalized["acceptanceEvents"] = json_value(normalized.get("acceptanceEvents"), [])
    if not isinstance(normalized["acceptanceEvents"], list):
        normalized["acceptanceEvents"] = []
    normalized["auditEvents"] = json_value(normalized.get("auditEvents"), [])
    if not isinstance(normalized["auditEvents"], list):
        normalized["auditEvents"] = []
    normalized["qualityIssueEvents"] = json_value(normalized.get("qualityIssueEvents"), [])
    if not isinstance(normalized["qualityIssueEvents"], list):
        normalized["qualityIssueEvents"] = []
    normalized.setdefault("qualityStatus", "pending")
    normalized["qualityFindings"] = json_value(normalized.get("qualityFindings"), [])
    if not isinstance(normalized["qualityFindings"], list):
        normalized["qualityFindings"] = []
    normalized.setdefault("reviewRecommendation", "")
    normalized.setdefault("publishRiskStatus", "unknown")
    normalized.setdefault("createdBy", "")
    normalized.setdefault("createdAt", None)
    normalized.setdefault("completedAt", None)
    normalized.setdefault("rolledBackAt", None)
    normalized.setdefault("deletedAt", None)
    return normalized


def load_import_reports_json() -> list[dict[str, Any]]:
    return [normalize_import_report_json(report) for report in load_json_records(import_batches_json_path())]


def context_has_import_batch_scope(context: AuthContext | None) -> bool:
    if context is None:
        return False
    return has_effective_area_scope(context) or any(
        has_effective_data_scope(context, scope_key)
        for scope_key, _api_field in IMPORT_BATCH_FINE_SCOPE_FIELDS
    )


def import_report_scope_records(report: dict[str, Any]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for item in report.get("importedBlocks") or []:
        if isinstance(item, dict):
            records.append(item)
    for archive in report.get("importedRightsArchives") or []:
        if not isinstance(archive, dict):
            continue
        linked_codes = archive.get("linkedBlockCodes") or []
        if isinstance(linked_codes, str):
            linked_codes = [linked_codes]
        for block_code in linked_codes:
            records.append({"blockCode": block_code})
    return records


def import_report_scope_record_allowed(context: AuthContext, record: dict[str, Any]) -> bool:
    block_code = str(record.get("blockCode") or "").strip()
    block = block_by_code(block_code, include_deleted=True) if block_code else None
    candidate = dict(block or {})
    candidate.update({key: value for key, value in record.items() if value not in (None, "")})
    if not candidate:
        return False
    if not area_allowed(context, candidate.get("countyCode")):
        return False
    return all(
        data_scope_value_allowed(context, scope_key, candidate.get(api_field))
        for scope_key, api_field in IMPORT_BATCH_FINE_SCOPE_FIELDS
    )


def import_report_allowed(context: AuthContext | None, report: dict[str, Any]) -> bool:
    if not context_has_import_batch_scope(context):
        return True
    assert context is not None
    if use_mysql() and report.get("_targetsLoaded") is False:
        return mysql_import_batch_visible(context, str(report.get("id") or ""))
    return any(
        import_report_scope_record_allowed(context, record)
        for record in import_report_scope_records(report)
    )


def require_import_report_allowed(context: AuthContext, report: dict[str, Any]) -> None:
    if not import_report_allowed(context, report):
        raise HTTPException(status_code=404, detail="Import batch not found")


def report_matches(
    report: dict[str, Any],
    filters: ImportBatchFilters,
    context: AuthContext | None = None,
) -> bool:
    if not import_report_allowed(context, report):
        return False
    if report.get("deletedAt") and not filters.includeDeleted:
        return False
    if filters.status and report.get("status") != filters.status:
        return False
    if filters.reviewStatus and str(report.get("reviewStatus") or "pending") != filters.reviewStatus:
        return False
    if filters.acceptanceStatus and str(report.get("acceptanceStatus") or "pending") != filters.acceptanceStatus:
        return False
    if filters.qualityStatus and str(report.get("qualityStatus") or "pending") != filters.qualityStatus:
        return False
    if filters.publishRiskStatus and str(report.get("publishRiskStatus") or "unknown") != filters.publishRiskStatus:
        return False
    if filters.sceneId:
        links = report.get("imageryLinks") or []
        if not any(str(item.get("sceneId") or "") == filters.sceneId for item in links if isinstance(item, dict)):
            return False
    if not import_report_matches_workflow_queue(report, filters.workflowQueue):
        return False
    if not filters.q:
        return True
    haystack = " ".join(
        [
            str(report.get("id") or ""),
            str(report.get("fileName") or ""),
            str(report.get("fileType") or ""),
            str(report.get("status") or ""),
            json.dumps(report.get("errors") or [], ensure_ascii=False),
            json.dumps(report.get("importedBlocks") or [], ensure_ascii=False),
            json.dumps(report.get("importedRightsArchives") or [], ensure_ascii=False),
        ]
    ).lower()
    return filters.q.lower() in haystack


def list_import_reports(filters: ImportBatchFilters, context: AuthContext | None = None) -> dict[str, Any]:
    if use_mysql():
        return {
            "items": list_import_reports_mysql(filters, context=context),
            "total": count_import_reports_mysql(filters, context),
            "limit": filters.limit,
            "offset": filters.offset,
        }
    if use_postgis():
        if context_has_import_batch_scope(context):
            reports = [
                report
                for report in list_import_reports_postgis(filters, paginate=False)
                if report_matches(report, filters, context)
            ]
            return {
                "items": reports[filters.offset : filters.offset + filters.limit],
                "total": len(reports),
                "limit": filters.limit,
                "offset": filters.offset,
            }
        return {
            "items": list_import_reports_postgis(filters),
            "total": count_import_reports_postgis(filters),
            "limit": filters.limit,
            "offset": filters.offset,
        }
    reports = [report for report in load_import_reports_json() if report_matches(report, filters, context)]
    return {
        "items": reports[filters.offset : filters.offset + filters.limit],
        "total": len(reports),
        "limit": filters.limit,
        "offset": filters.offset,
    }


def audit_event_matches(event: dict[str, Any], *, q: str, action: str, batch_id: str) -> bool:
    if action and str(event.get("action") or "") != action:
        return False
    if batch_id and str(event.get("batchId") or "") != batch_id:
        return False
    if not q:
        return True
    haystack = " ".join(
        [
            str(event.get("batchId") or ""),
            str(event.get("fileName") or ""),
            str(event.get("action") or ""),
            str(event.get("actor") or ""),
            json.dumps(event.get("summary") or {}, ensure_ascii=False),
        ]
    ).lower()
    return q.lower() in haystack


def import_audit_event_record(report: dict[str, Any], event: dict[str, Any], index: int) -> dict[str, Any]:
    batch_id = str(report.get("id") or "")
    return {
        "eventId": f"{batch_id}:{index}",
        "batchId": batch_id,
        "fileName": report.get("fileName") or "",
        "batchStatus": report.get("status") or "",
        "reviewStatus": report.get("reviewStatus") or "pending",
        "acceptanceStatus": report.get("acceptanceStatus") or "pending",
        "qualityStatus": report.get("qualityStatus") or "pending",
        "publishRiskStatus": report.get("publishRiskStatus") or "unknown",
        "at": event.get("at"),
        "action": event.get("action") or "",
        "actor": event.get("actor") or "",
        "summary": event.get("summary") or {},
    }


def list_import_audit_events_mysql(
    *,
    q: str = "",
    action: str = "",
    batch_id: str = "",
    include_deleted: bool = False,
    limit: int = 100,
    offset: int = 0,
) -> dict[str, Any]:
    clauses = ["ibe.event_type = 'audit'"]
    params: list[Any] = []
    if not include_deleted:
        clauses.append("ib.deleted_at IS NULL")
    if action:
        clauses.append("ibe.action = %s")
        params.append(action)
    if batch_id:
        clauses.append("ibe.import_batch_id = %s")
        params.append(batch_id)
    if q.strip():
        like = f"%{q.strip()}%"
        clauses.append(
            "(ib.file_name LIKE %s OR ibe.action LIKE %s OR ibe.actor LIKE %s "
            "OR ibe.summary LIKE %s OR CAST(ibe.event_json AS CHAR) LIKE %s)"
        )
        params.extend([like] * 5)
    where_sql = " WHERE " + " AND ".join(clauses)
    from_sql = " FROM import_batch_events ibe JOIN import_batches ib ON ib.id = ibe.import_batch_id"
    with mysql_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*){from_sql}{where_sql}", tuple(params))
            count_row = cur.fetchone()
            total = int(count_row[0] if count_row else 0)
            cur.execute(
                "SELECT ibe.event_id, ibe.import_batch_id, ib.file_name, ib.status, "
                "ib.review_status, ib.acceptance_status, ib.quality_status, ib.publish_risk_status, "
                "ibe.event_at, ibe.action, ibe.actor, ibe.event_json"
                f"{from_sql}{where_sql} ORDER BY ibe.event_at DESC, ibe.event_id DESC LIMIT %s OFFSET %s",
                tuple([*params, limit, offset]),
            )
            rows = cur.fetchall()
    items: list[dict[str, Any]] = []
    for row in rows:
        event = json_value(row[11], {})
        items.append(
            {
                "eventId": row[0],
                "batchId": row[1],
                "fileName": row[2] or "",
                "batchStatus": row[3] or "",
                "reviewStatus": row[4] or "pending",
                "acceptanceStatus": row[5] or "pending",
                "qualityStatus": row[6] or "pending",
                "publishRiskStatus": row[7] or "unknown",
                "at": datetime_to_iso(row[8]),
                "action": row[9] or "",
                "actor": row[10] or "",
                "summary": event.get("summary") if isinstance(event, dict) else {},
            }
        )
    return {"items": items, "total": total, "limit": limit, "offset": offset}


def list_import_audit_events(
    *,
    q: str = "",
    action: str = "",
    batch_id: str = "",
    include_deleted: bool = False,
    limit: int = 100,
    offset: int = 0,
    context: AuthContext | None = None,
) -> dict[str, Any]:
    if use_mysql() and not context_has_import_batch_scope(context):
        return list_import_audit_events_mysql(
            q=q,
            action=action,
            batch_id=batch_id,
            include_deleted=include_deleted,
            limit=limit,
            offset=offset,
        )
    report_payload = list_import_reports(
        ImportBatchFilters(includeDeleted=include_deleted, limit=1000, offset=0),
        context,
    )
    events: list[dict[str, Any]] = []
    for report in report_payload.get("items") or []:
        for index, event in enumerate(report.get("auditEvents") or []):
            if not isinstance(event, dict):
                continue
            record = import_audit_event_record(report, event, index)
            if audit_event_matches(record, q=q, action=action, batch_id=batch_id):
                events.append(record)
    events.sort(key=lambda item: str(item.get("at") or ""), reverse=True)
    return {
        "items": events[offset : offset + limit],
        "total": len(events),
        "limit": limit,
        "offset": offset,
    }


def csv_download_response(filename: str, columns: list[str], records: list[dict[str, Any]]) -> Response:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(columns)
    for record in records:
        writer.writerow(
            [
                json.dumps(record.get(column), ensure_ascii=False, sort_keys=True)
                if isinstance(record.get(column), (dict, list))
                else record.get(column, "")
                for column in columns
            ]
        )
    content = "\ufeff" + output.getvalue()
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def import_quality_issue_record(
    report: dict[str, Any],
    *,
    issue_type: str,
    issue_key: str,
    severity: str,
    message: str,
    action_required: str,
    block_codes: list[str] | None = None,
    scene_id: str = "",
    source: str = "batch",
) -> dict[str, Any]:
    batch_id = str(report.get("id") or "")
    safe_key = re.sub(r"[^a-zA-Z0-9_.:-]+", "-", issue_key or issue_type).strip("-")
    scene_suffix = f":{scene_id}" if scene_id else ""
    return {
        "issueId": f"{batch_id}:{issue_type}:{safe_key}{scene_suffix}",
        "batchId": batch_id,
        "fileName": report.get("fileName") or "",
        "batchStatus": report.get("status") or "",
        "reviewStatus": report.get("reviewStatus") or "pending",
        "qualityStatus": report.get("qualityStatus") or "pending",
        "publishRiskStatus": report.get("publishRiskStatus") or "unknown",
        "issueType": issue_type,
        "issueKey": issue_key,
        "severity": severity,
        "status": "open",
        "message": message,
        "actionRequired": action_required,
        "blockCodes": unique_values(block_codes or []),
        "sceneId": scene_id,
        "source": source,
        "createdAt": report.get("completedAt") or report.get("createdAt"),
        "updatedAt": report.get("updatedAt") or report.get("completedAt") or report.get("createdAt"),
    }


def latest_quality_issue_event(report: dict[str, Any], issue_id: str) -> dict[str, Any] | None:
    for event in reversed(report.get("qualityIssueEvents") or []):
        if isinstance(event, dict) and str(event.get("issueId") or "") == issue_id:
            return event
    return None


def apply_quality_issue_event(report: dict[str, Any], issue: dict[str, Any]) -> dict[str, Any]:
    event = latest_quality_issue_event(report, str(issue.get("issueId") or ""))
    if not event:
        return issue
    updated = dict(issue)
    updated["status"] = event.get("status") or updated.get("status") or "open"
    updated["handledBy"] = event.get("actor") or ""
    updated["handledAt"] = event.get("at")
    updated["handlingComment"] = event.get("comment") or ""
    updated["updatedAt"] = event.get("at") or updated.get("updatedAt")
    return updated


def coverage_warning_block_codes(coverage_check: dict[str, Any], warning: str) -> list[str]:
    if warning == "missing_geometry":
        return unique_values(coverage_check.get("missingGeometryBlockCodes") or [])
    if warning == "outside_scene_bounds":
        return unique_values(coverage_check.get("outsideSceneBoundsBlockCodes") or [])
    return []


def quality_issues_for_report(report: dict[str, Any]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for index, error in enumerate(report.get("errors") or [], start=1):
        if not isinstance(error, dict):
            continue
        row = str(error.get("row") or index)
        issues.append(
            import_quality_issue_record(
                report,
                issue_type="import_error",
                issue_key=f"row_{row}",
                severity="blocked",
                message=str(error.get("message") or "Import row error"),
                action_required="修正源文件错误行后重新导入",
                source="import",
            )
        )

    review_status = str(report.get("reviewStatus") or "pending")
    if review_status in {"needs_correction", "rejected"}:
        issues.append(
            import_quality_issue_record(
                report,
                issue_type="review_decision",
                issue_key=review_status,
                severity="blocked" if review_status == "rejected" else "warning",
                message=str(report.get("reviewComment") or review_status),
                action_required="按审核意见修正后重新提交审核",
                source="review",
            )
        )

    missing_block_codes = [
        block_code
        for block_code in import_report_block_codes(report)
        if not block_by_code(block_code)
    ]
    if missing_block_codes:
        issues.append(
            import_quality_issue_record(
                report,
                issue_type="missing_imported_block",
                issue_key="block_not_found",
                severity="blocked",
                message="导入批次中的部分林班在当前空间台账中不存在或已删除",
                action_required="恢复或重新导入缺失林班后再发布图层",
                block_codes=missing_block_codes,
                source="ledger",
            )
        )

    risk_status = str(report.get("publishRiskStatus") or "unknown")
    quality_severity = "blocked" if risk_status == "blocked" else "warning"
    for finding in unique_values(report.get("qualityFindings") or []):
        issues.append(
            import_quality_issue_record(
                report,
                issue_type="quality_finding",
                issue_key=finding,
                severity=quality_severity,
                message=finding,
                action_required="处理质量发现并重新执行发布预检",
                source="quality",
            )
        )

    for link in report.get("imageryLinks") or []:
        if not isinstance(link, dict):
            continue
        coverage_check = link.get("coverageCheck") or {}
        if not isinstance(coverage_check, dict):
            continue
        scene_id = str(link.get("sceneId") or "")
        for warning in unique_values(coverage_check.get("warnings") or []):
            issues.append(
                import_quality_issue_record(
                    report,
                    issue_type="coverage_warning",
                    issue_key=warning,
                    severity="warning",
                    message=warning,
                    action_required="核验影像范围、林班边界和图层发布范围",
                    block_codes=coverage_warning_block_codes(coverage_check, warning),
                    scene_id=scene_id,
                    source="coverage",
                )
            )

    return [apply_quality_issue_event(report, issue) for issue in issues]


def quality_issue_matches(
    issue: dict[str, Any],
    *,
    q: str,
    issue_type: str,
    severity: str,
    batch_id: str,
    status: str,
) -> bool:
    if issue_type and str(issue.get("issueType") or "") != issue_type:
        return False
    if severity and str(issue.get("severity") or "") != severity:
        return False
    if batch_id and str(issue.get("batchId") or "") != batch_id:
        return False
    if status and str(issue.get("status") or "open") != status:
        return False
    if not q:
        return True
    haystack = " ".join(
        [
            str(issue.get("issueId") or ""),
            str(issue.get("batchId") or ""),
            str(issue.get("fileName") or ""),
            str(issue.get("issueType") or ""),
            str(issue.get("issueKey") or ""),
            str(issue.get("severity") or ""),
            str(issue.get("message") or ""),
            str(issue.get("actionRequired") or ""),
            str(issue.get("sceneId") or ""),
            " ".join(issue.get("blockCodes") or []),
        ]
    ).lower()
    return q.lower() in haystack


def list_import_quality_issues(
    *,
    q: str = "",
    issue_type: str = "",
    severity: str = "",
    batch_id: str = "",
    status: str = "",
    include_deleted: bool = False,
    limit: int = 100,
    offset: int = 0,
    context: AuthContext | None = None,
) -> dict[str, Any]:
    report_payload = list_import_reports(
        ImportBatchFilters(includeDeleted=include_deleted, limit=1000, offset=0),
        context,
    )
    issues: list[dict[str, Any]] = []
    for report in report_payload.get("items") or []:
        for issue in quality_issues_for_report(report):
            if quality_issue_matches(issue, q=q, issue_type=issue_type, severity=severity, batch_id=batch_id, status=status):
                issues.append(issue)
    issues.sort(
        key=lambda item: (
            {"blocked": 0, "warning": 1, "info": 2}.get(str(item.get("severity") or ""), 3),
            str(item.get("updatedAt") or ""),
            str(item.get("issueId") or ""),
        ),
    )
    return {
        "items": issues[offset : offset + limit],
        "total": len(issues),
        "limit": limit,
        "offset": offset,
    }


def workflow_summary_card(key: str, label: str, value: int, tone: str, href: str) -> dict[str, Any]:
    return {"key": key, "label": label, "value": value, "tone": tone, "href": href}


def active_import_workflow_reports(context: AuthContext) -> list[dict[str, Any]]:
    payload = list_import_reports(
        ImportBatchFilters(includeDeleted=False, limit=1000, offset=0),
        context,
    )
    return [
        report
        for report in payload.get("items") or []
        if not report.get("deletedAt") and str(report.get("status") or "") not in {"deleted", "rolled_back"}
    ]


def import_report_has_imagery_link(report: dict[str, Any]) -> bool:
    return any(
        isinstance(link, dict) and str(link.get("sceneId") or "").strip()
        for link in report.get("imageryLinks") or []
    )


def import_report_ready_for_layer_link(report: dict[str, Any]) -> bool:
    risk_status = str(report.get("publishRiskStatus") or "unknown")
    return (
        str(report.get("reviewStatus") or "pending") == "approved"
        and not import_report_has_imagery_link(report)
        and int(report.get("validRows") or 0) > 0
        and str(report.get("qualityStatus") or "pending") in {"passed", "warning"}
        and risk_status in {"clear", "ready", "warning"}
    )


def import_report_matches_workflow_queue(report: dict[str, Any], workflow_queue: str) -> bool:
    queue = workflow_queue.strip()
    if not queue:
        return True
    if queue == "pendingReview":
        return str(report.get("reviewStatus") or "pending") == "pending"
    if queue == "approvedUnlinked":
        return str(report.get("reviewStatus") or "pending") == "approved" and not import_report_has_imagery_link(report)
    if queue == "readyForLayerLink":
        return import_report_ready_for_layer_link(report)
    if queue == "linked":
        return import_report_has_imagery_link(report)
    return False


def import_delivery_layer_for_scene(scene: dict[str, Any] | None, scene_id: str, link: dict[str, Any]) -> dict[str, Any]:
    scene = scene or {}
    published_layer_id = str(scene.get("publishedLayerId") or link.get("layerId") or "")
    published_record_code = str(scene.get("publishedLayerRecordCode") or link.get("layerRecordCode") or "")
    fallback_record_code = f"SCENE-LAYER-{scene_id}"
    for record in layer_records():
        properties = record.get("properties") if isinstance(record.get("properties"), dict) else {}
        if (
            (published_layer_id and str(record.get("id")) == published_layer_id)
            or (published_record_code and str(record.get("recordCode")) == published_record_code)
            or str(record.get("recordCode")) == fallback_record_code
            or str(properties.get("sourceSceneId") or "") == scene_id
            or str(record.get("dataSource") or "") == f"scene:{scene_id}"
        ):
            return enrich_map_layer_record(record)
    return {}


def import_delivery_unique_scene_links(report: dict[str, Any]) -> list[dict[str, Any]]:
    links_by_scene: dict[str, dict[str, Any]] = {}
    for link in report.get("imageryLinks") or []:
        if not isinstance(link, dict):
            continue
        scene_id = str(link.get("sceneId") or "").strip()
        if scene_id:
            links_by_scene[scene_id] = dict(link)
    return list(links_by_scene.values())


def import_delivery_scene_item(link: dict[str, Any], context: AuthContext) -> dict[str, Any]:
    scene_id = str(link.get("sceneId") or "").strip()
    scene: dict[str, Any] | None = None
    missing = False
    try:
        scene = require_visible_scene(scene_id, context)
    except HTTPException as exc:
        if exc.status_code not in {403, 404}:
            raise
        missing = True

    layer = import_delivery_layer_for_scene(scene, scene_id, link)
    delivery_status = str((scene or {}).get("deliveryStatus") or "pending")
    scene_status = str((scene or {}).get("status") or "")
    published = bool(layer) or bool((scene or {}).get("publishedLayerId")) or bool(
        (scene or {}).get("publishedLayerRecordCode")
    ) or scene_status == "published"
    return {
        "sceneId": scene_id,
        "sceneName": (scene or {}).get("name") or link.get("sceneName") or scene_id,
        "status": scene_status or ("missing" if missing else "active"),
        "deliveryStatus": delivery_status,
        "deliveryReceiptUrl": f"/api/scenes/{scene_id}/delivery-receipt.json" if scene_id and not missing else "",
        "deliveredAt": (scene or {}).get("deliveredAt"),
        "deliveredBy": (scene or {}).get("deliveredBy") or "",
        "published": published,
        "sceneMissing": missing,
        "publishedLayerId": layer.get("id") or (scene or {}).get("publishedLayerId") or link.get("layerId") or "",
        "publishedLayerRecordCode": layer.get("recordCode")
        or (scene or {}).get("publishedLayerRecordCode")
        or link.get("layerRecordCode")
        or "",
        "adminHref": f"admin-imagery.html?sceneId={scene_id}",
    }


def import_delivery_aggregate_status(scenes: list[dict[str, Any]]) -> str:
    statuses = [str(scene.get("deliveryStatus") or "pending") for scene in scenes]
    if not statuses:
        return "pending"
    if any(status == "rejected" for status in statuses):
        return "rejected"
    if any(status == "needs_correction" for status in statuses):
        return "needs_correction"
    if all(status == "delivered" for status in statuses):
        return "delivered"
    if any(status == "delivered" for status in statuses):
        return "partial"
    return "pending"


def import_delivery_package_status(
    report: dict[str, Any],
    scenes: list[dict[str, Any]],
    blocking_reasons: list[str],
    delivery_status: str,
) -> str:
    if any(reason in {"quality_blocked", "publish_risk_blocked", "scene_missing"} for reason in blocking_reasons):
        return "blocked"
    if str(report.get("reviewStatus") or "pending") != "approved":
        return "awaiting_review"
    if str(report.get("acceptanceStatus") or "pending") != "accepted":
        return "awaiting_acceptance"
    if not scenes:
        return "awaiting_scene_link"
    if any(not scene.get("published") for scene in scenes):
        return "awaiting_publish"
    if delivery_status != "delivered":
        return "awaiting_delivery"
    return "ready"


def import_delivery_blocking_reasons(report: dict[str, Any], scenes: list[dict[str, Any]]) -> list[str]:
    reasons: list[str] = []
    if str(report.get("reviewStatus") or "pending") != "approved":
        reasons.append("review_not_approved")
    if str(report.get("acceptanceStatus") or "pending") != "accepted":
        reasons.append("acceptance_not_accepted")
    if not scenes:
        reasons.append("no_linked_scene")
    if str(report.get("qualityStatus") or "pending") == "blocked":
        reasons.append("quality_blocked")
    if str(report.get("publishRiskStatus") or "unknown") == "blocked":
        reasons.append("publish_risk_blocked")
    for scene in scenes:
        if scene.get("sceneMissing"):
            reasons.append("scene_missing")
        if not scene.get("published"):
            reasons.append("scene_not_published")
        if str(scene.get("deliveryStatus") or "pending") != "delivered":
            reasons.append("scene_not_delivered")
    return unique_values(reasons)


def import_delivery_map_layer_receipt_urls(record_codes: list[str]) -> list[dict[str, str]]:
    if not record_codes:
        return []
    wanted_codes = {str(code or "") for code in record_codes if str(code or "")}
    layers_by_code = {
        str(layer.get("recordCode") or ""): layer
        for layer in layer_records()
        if str(layer.get("recordCode") or "") in wanted_codes
    }
    receipts: list[dict[str, str]] = []
    for record_code in record_codes:
        layer = layers_by_code.get(str(record_code or ""))
        layer_id = str(layer.get("id") or "") if layer else ""
        if not layer_id:
            continue
        receipts.append(
            {
                "recordCode": str(record_code or ""),
                "layerId": layer_id,
                "url": f"/api/map-layers/{layer_id}/publication-receipt.json",
            }
        )
    return receipts


def import_delivery_package_for_report(report: dict[str, Any], context: AuthContext) -> dict[str, Any]:
    scene_links = import_delivery_unique_scene_links(report)
    scenes = [import_delivery_scene_item(link, context) for link in scene_links]
    delivery_status = import_delivery_aggregate_status(scenes)
    blocking_reasons = import_delivery_blocking_reasons(report, scenes)
    package_status = import_delivery_package_status(report, scenes, blocking_reasons, delivery_status)
    linked_block_codes = import_report_block_codes(report)
    published_layer_codes = unique_values(
        [scene.get("publishedLayerRecordCode") for scene in scenes if scene.get("published")]
    )
    map_layer_publication_receipt_urls = import_delivery_map_layer_receipt_urls(published_layer_codes)
    receipt_scenes = [
        scene
        for scene in scenes
        if scene.get("sceneId") and not scene.get("sceneMissing") and scene.get("deliveryReceiptUrl")
    ]
    primary_scene = receipt_scenes[0] if receipt_scenes else (scenes[0] if scenes else {})
    primary_scene_id = str(primary_scene.get("sceneId") or "")
    primary_scene_delivery_receipt_url = str(primary_scene.get("deliveryReceiptUrl") or "")
    scene_delivery_receipt_urls = [
        {"sceneId": str(scene.get("sceneId") or ""), "url": str(scene.get("deliveryReceiptUrl") or "")}
        for scene in receipt_scenes
    ]
    return {
        "batchId": report.get("id") or "",
        "fileName": report.get("fileName") or "",
        "status": report.get("status") or "",
        "reviewStatus": report.get("reviewStatus") or "pending",
        "acceptanceStatus": report.get("acceptanceStatus") or "pending",
        "acceptedAt": report.get("acceptedAt"),
        "acceptedBy": report.get("acceptedBy") or "",
        "qualityStatus": report.get("qualityStatus") or "pending",
        "publishRiskStatus": report.get("publishRiskStatus") or "unknown",
        "packageStatus": package_status,
        "deliveryStatus": delivery_status,
        "linkedBlockCodes": linked_block_codes,
        "linkedBlockCount": len(linked_block_codes),
        "rightArchiveCount": len(import_report_right_archive_codes(report)),
        "linkedSceneCount": len(scenes),
        "deliveredSceneCount": len(
            [scene for scene in scenes if str(scene.get("deliveryStatus") or "") == "delivered"]
        ),
        "pendingSceneCount": len(
            [scene for scene in scenes if str(scene.get("deliveryStatus") or "pending") != "delivered"]
        ),
        "publishedLayerCount": len(published_layer_codes),
        "publishedLayerRecordCodes": published_layer_codes,
        "mapLayerPublicationReceiptUrls": map_layer_publication_receipt_urls,
        "blockingReasons": blocking_reasons,
        "scenes": scenes,
        "acceptanceReceiptUrl": f"/api/imports/{report.get('id') or ''}/acceptance-receipt.json",
        "primarySceneId": primary_scene_id,
        "primarySceneDeliveryReceiptUrl": primary_scene_delivery_receipt_url,
        "sceneDeliveryReceiptUrls": scene_delivery_receipt_urls,
        "updatedAt": report.get("completedAt") or report.get("createdAt") or "",
        "adminHref": f"admin-imports.html?batchId={report.get('id') or ''}",
    }


def import_delivery_package_matches(
    package: dict[str, Any],
    *,
    q: str,
    status: str,
    acceptance_status: str,
    delivery_status: str,
    linked_block_code: str,
) -> bool:
    if status and str(package.get("packageStatus") or "") != status:
        return False
    if acceptance_status and str(package.get("acceptanceStatus") or "") != acceptance_status:
        return False
    if delivery_status and str(package.get("deliveryStatus") or "") != delivery_status:
        return False
    if linked_block_code and linked_block_code not in set(package.get("linkedBlockCodes") or []):
        return False
    if not q:
        return True
    haystack = " ".join(
        [
            str(package.get("batchId") or ""),
            str(package.get("fileName") or ""),
            str(package.get("packageStatus") or ""),
            str(package.get("deliveryStatus") or ""),
            json.dumps(package.get("blockingReasons") or [], ensure_ascii=False),
            json.dumps(package.get("linkedBlockCodes") or [], ensure_ascii=False),
            json.dumps(package.get("scenes") or [], ensure_ascii=False),
        ]
    ).lower()
    return q.lower() in haystack


def import_delivery_package_summary(packages: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "packageTotal": len(packages),
        "readyPackages": len([item for item in packages if item.get("packageStatus") == "ready"]),
        "awaitingReviewPackages": len(
            [item for item in packages if item.get("packageStatus") == "awaiting_review"]
        ),
        "awaitingAcceptancePackages": len(
            [item for item in packages if item.get("packageStatus") == "awaiting_acceptance"]
        ),
        "awaitingSceneLinkPackages": len(
            [item for item in packages if item.get("packageStatus") == "awaiting_scene_link"]
        ),
        "awaitingPublishPackages": len(
            [item for item in packages if item.get("packageStatus") == "awaiting_publish"]
        ),
        "awaitingDeliveryPackages": len(
            [item for item in packages if item.get("packageStatus") == "awaiting_delivery"]
        ),
        "blockedPackages": len([item for item in packages if item.get("packageStatus") == "blocked"]),
    }


def import_delivery_packages_receipt(payload: dict[str, Any], context: AuthContext) -> dict[str, Any]:
    return {
        "receiptType": "import-delivery-packages",
        "exportedAt": now_iso(),
        **import_receipt_export_metadata(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION),
        "summary": payload.get("summary") or {},
        "items": payload.get("items") or [],
        "total": payload.get("total") or 0,
        "limit": payload.get("limit") or 0,
        "offset": payload.get("offset") or 0,
    }


def import_delivery_package_receipt(report: dict[str, Any], context: AuthContext) -> dict[str, Any]:
    package = import_delivery_package_for_report(report, context)
    blocking_reasons = package.get("blockingReasons") or []
    scenes = package.get("scenes") or []
    return {
        "receiptType": "import-delivery-package",
        "exportedAt": now_iso(),
        **import_receipt_export_metadata(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION),
        "batch": report,
        "deliveryPackage": package,
        "summary": {
            "batchId": package.get("batchId") or "",
            "fileName": package.get("fileName") or "",
            "packageStatus": package.get("packageStatus") or "",
            "readyForDelivery": package.get("packageStatus") == "ready",
            "acceptanceStatus": package.get("acceptanceStatus") or "pending",
            "deliveryStatus": package.get("deliveryStatus") or "pending",
            "qualityStatus": package.get("qualityStatus") or "pending",
            "publishRiskStatus": package.get("publishRiskStatus") or "unknown",
            "linkedBlockCount": package.get("linkedBlockCount") or 0,
            "linkedSceneCount": package.get("linkedSceneCount") or 0,
            "deliveredSceneCount": package.get("deliveredSceneCount") or 0,
            "publishedLayerCount": package.get("publishedLayerCount") or 0,
            "blockingReasonCount": len(blocking_reasons),
            "receiptCount": 1
            + len(package.get("sceneDeliveryReceiptUrls") or [])
            + len(package.get("mapLayerPublicationReceiptUrls") or []),
        },
        "acceptanceReceiptUrl": package.get("acceptanceReceiptUrl") or "",
        "sceneDeliveryReceiptUrls": package.get("sceneDeliveryReceiptUrls") or [],
        "mapLayerPublicationReceiptUrls": package.get("mapLayerPublicationReceiptUrls") or [],
        "scenes": scenes,
        "blockingReasons": blocking_reasons,
        "auditEvents": list(report.get("auditEvents") or []),
    }


def list_import_delivery_packages(
    *,
    q: str = "",
    status: str = "",
    acceptance_status: str = "",
    delivery_status: str = "",
    linked_block_code: str = "",
    limit: int = 100,
    offset: int = 0,
    context: AuthContext,
) -> dict[str, Any]:
    if status and status not in IMPORT_DELIVERY_PACKAGE_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"status must be one of: {', '.join(sorted(IMPORT_DELIVERY_PACKAGE_STATUSES))}",
        )
    if acceptance_status and acceptance_status not in IMPORT_ACCEPTANCE_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"acceptanceStatus must be one of: {', '.join(sorted(IMPORT_ACCEPTANCE_STATUSES))}",
        )
    if delivery_status and delivery_status not in IMPORT_DELIVERY_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"deliveryStatus must be one of: {', '.join(sorted(IMPORT_DELIVERY_STATUSES))}",
        )
    reports = active_import_workflow_reports(context)
    packages = [import_delivery_package_for_report(report, context) for report in reports]
    packages = [
        package
        for package in packages
        if import_delivery_package_matches(
            package,
            q=q.strip(),
            status=status.strip(),
            acceptance_status=acceptance_status.strip(),
            delivery_status=delivery_status.strip(),
            linked_block_code=linked_block_code.strip(),
        )
    ]
    packages.sort(key=lambda item: (str(item.get("updatedAt") or ""), str(item.get("batchId") or "")), reverse=True)
    return {
        "items": packages[offset : offset + limit],
        "total": len(packages),
        "limit": limit,
        "offset": offset,
        "summary": import_delivery_package_summary(packages),
    }


def import_operation_queue_item(package: dict[str, Any], stage: dict[str, Any]) -> dict[str, Any]:
    batch_id = str(package.get("batchId") or "")
    scenes = package.get("scenes") if isinstance(package.get("scenes"), list) else []
    primary_scene = scenes[0] if scenes else {}
    return {
        "batchId": batch_id,
        "fileName": package.get("fileName") or "",
        "packageStatus": package.get("packageStatus") or "",
        "reviewStatus": package.get("reviewStatus") or "pending",
        "acceptanceStatus": package.get("acceptanceStatus") or "pending",
        "qualityStatus": package.get("qualityStatus") or "pending",
        "publishRiskStatus": package.get("publishRiskStatus") or "unknown",
        "deliveryStatus": package.get("deliveryStatus") or "pending",
        "linkedBlockCount": package.get("linkedBlockCount") or 0,
        "linkedSceneCount": package.get("linkedSceneCount") or 0,
        "publishedLayerCount": package.get("publishedLayerCount") or 0,
        "pendingSceneCount": package.get("pendingSceneCount") or 0,
        "blockingReasons": package.get("blockingReasons") or [],
        "primarySceneId": package.get("primarySceneId") or primary_scene.get("sceneId") or "",
        "primaryAction": stage.get("primaryAction") or "",
        "primaryActionLabel": stage.get("primaryActionLabel") or "",
        "requiredPermission": stage.get("requiredPermission") or "",
        "requiredAllPermissions": stage.get("requiredAllPermissions") or [],
        "requiredAnyPermissions": stage.get("requiredAnyPermissions") or [],
        "adminHref": package.get("adminHref") or f"admin-imports.html?batchId={batch_id}",
        "sceneAdminHref": primary_scene.get("adminHref") or "",
        "updatedAt": package.get("updatedAt") or "",
    }


def import_operation_queue(*, limit: int, context: AuthContext) -> dict[str, Any]:
    payload = list_import_delivery_packages(limit=1000, offset=0, context=context)
    packages = payload.get("items") or []
    lanes: list[dict[str, Any]] = []
    for stage in IMPORT_OPERATION_QUEUE_STAGES:
        status = str(stage.get("key") or "")
        matches = [package for package in packages if str(package.get("packageStatus") or "") == status]
        lanes.append(
            {
                **stage,
                "count": len(matches),
                "items": [import_operation_queue_item(package, stage) for package in matches[:limit]],
            }
        )
    package_summary = import_delivery_package_summary(packages)
    actionable_count = sum(
        [
            package_summary.get("awaitingReviewPackages", 0),
            package_summary.get("awaitingAcceptancePackages", 0),
            package_summary.get("awaitingSceneLinkPackages", 0),
            package_summary.get("awaitingPublishPackages", 0),
            package_summary.get("awaitingDeliveryPackages", 0),
        ]
    )
    blocked_count = package_summary.get("blockedPackages", 0)
    return {
        "items": lanes,
        "total": len(lanes),
        "limit": limit,
        "generatedAt": now_iso(),
        "summary": {
            **package_summary,
            "operationQueueTotal": actionable_count + blocked_count,
            "actionableQueueTotal": actionable_count,
            "blockedQueueTotal": blocked_count,
            "readyQueueTotal": package_summary.get("readyPackages", 0),
        },
    }


def import_workflow_summary(context: AuthContext) -> dict[str, Any]:
    reports = active_import_workflow_reports(context)
    issues = [
        issue
        for issue in list_import_quality_issues(limit=1000, offset=0, context=context).get("items") or []
        if str(issue.get("status") or "open") not in {"resolved", "ignored"}
    ]
    pending_review = [
        report for report in reports if str(report.get("reviewStatus") or "pending") == "pending"
    ]
    approved_unlinked = [
        report
        for report in reports
        if str(report.get("reviewStatus") or "pending") == "approved"
        and not import_report_has_imagery_link(report)
    ]
    linked = [report for report in reports if import_report_has_imagery_link(report)]
    ready_for_layer_link = [report for report in reports if import_report_ready_for_layer_link(report)]
    delivery_packages = [import_delivery_package_for_report(report, context) for report in reports]
    ready_delivery_packages = [
        package for package in delivery_packages if str(package.get("packageStatus") or "") == "ready"
    ]
    blocked_issues = [issue for issue in issues if str(issue.get("severity") or "") == "blocked"]
    needs_attention = len(pending_review) + len(approved_unlinked) + len(issues)
    return {
        "activeBatchTotal": len(reports),
        "pendingReviewBatches": len(pending_review),
        "approvedUnlinkedBatches": len(approved_unlinked),
        "linkedBatches": len(linked),
        "readyForLayerLinkBatches": len(ready_for_layer_link),
        "readyDeliveryPackages": len(ready_delivery_packages),
        "openQualityIssues": len(issues),
        "blockedQualityIssues": len(blocked_issues),
        "needsAttentionTotal": needs_attention,
        "cards": [
            workflow_summary_card(
                "pendingReviewBatches",
                "待审核批次",
                len(pending_review),
                "warning",
                "admin-imports.html?reviewStatus=pending",
            ),
            workflow_summary_card(
                "approvedUnlinkedBatches",
                "已审未挂影像",
                len(approved_unlinked),
                "warning",
                "admin-imports.html?workflowQueue=approvedUnlinked",
            ),
            workflow_summary_card(
                "readyForLayerLinkBatches",
                "可关联发布",
                len(ready_for_layer_link),
                "ready",
                "admin-imports.html?workflowQueue=readyForLayerLink",
            ),
            workflow_summary_card(
                "readyDeliveryPackages",
                "可交付批次",
                len(ready_delivery_packages),
                "ready",
                "admin-imports.html?deliveryPackageStatus=ready",
            ),
            workflow_summary_card(
                "blockedQualityIssues",
                "阻断质检问题",
                len(blocked_issues),
                "danger",
                "admin-imports.html?qualityIssueStatus=open",
            ),
        ],
    }


def save_import_report_json(report: dict[str, Any]) -> None:
    reports = [item for item in load_import_reports_json() if str(item.get("id")) != str(report.get("id"))]
    reports.insert(0, report)
    save_json_records(import_batches_json_path(), reports)


def load_import_report_json(batch_id: str) -> dict[str, Any] | None:
    for report in load_import_reports_json():
        if str(report.get("id")) == str(batch_id):
            return report
    return None


def save_import_report(report: dict[str, Any]) -> None:
    batch_id = str(report["id"])
    if use_mysql():
        upsert_import_report_mysql(report)
        IMPORT_REPORTS.pop(batch_id, None)
        return
    if use_postgis():
        upsert_import_report_postgis(report)
        IMPORT_REPORTS.pop(batch_id, None)
        return
    IMPORT_REPORTS[batch_id] = report
    save_import_report_json(report)


def append_import_batch_audit_event(
    report: dict[str, Any],
    *,
    action: str,
    actor: str = "",
    summary: dict[str, Any] | None = None,
    timestamp: str | None = None,
) -> dict[str, Any]:
    event = {
        "at": timestamp or now_iso(),
        "action": action,
        "actor": actor,
        "summary": summary or {},
    }
    events = list(report.get("auditEvents") or [])
    events.append(event)
    report["auditEvents"] = events
    report["updatedAt"] = event["at"]
    return event


def execute_forest_block_import(
    *,
    records: list[dict[str, Any]],
    file_name: str,
    strategy: str,
    context: Any,
) -> dict[str, Any]:
    if strategy not in {"upsert", "skip"}:
        raise HTTPException(status_code=400, detail="strategy must be 'upsert' or 'skip'")

    database_backed = use_mysql() or use_postgis()
    blocks: list[dict[str, Any]] = []
    active_indexes: dict[str, int] = {}
    existing_identities: dict[str, dict[str, Any]] = {}
    if database_backed:
        existing_identities = block_identities_by_codes(
            [str(record.get("blockCode") or "") for record in records]
        )
    else:
        blocks = load_all_blocks()
        active_indexes = {
            block["blockCode"]: index
            for index, block in enumerate(blocks)
            if block.get("blockCode") and not block.get("deletedAt")
        }

    valid_rows = 0
    did_change_blocks = False
    pending_database_blocks: list[dict[str, Any]] = []
    pending_right_blocks: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    imported_blocks: list[dict[str, Any]] = []
    imported_rights_archives: list[dict[str, Any]] = []
    batch_id = str(uuid.uuid4())
    for row_number, record in enumerate(records, start=1):
        row_errors = validate_record(record)
        try:
            require_target_area_allowed(context, record.get("countyCode"))
        except HTTPException as exc:
            row_errors.append(str(exc.detail))
        if row_errors:
            errors.extend({"row": row_number, "message": message} for message in row_errors)
            continue

        valid_rows += 1
        clean_record = clean_import_record(record)
        normalized = normalize_block(clean_record)
        normalized["sourceBatchId"] = batch_id
        block_code = str(normalized.get("blockCode") or clean_record.get("blockCode") or "")
        block_name = str(normalized.get("name") or clean_record.get("name") or "")
        if is_rights_archive_like_block(normalized):
            pending_right_blocks.append(normalized)
            imported_rights_archives.append(
                {
                    "archiveCode": block_code,
                    "linkedBlockCodes": [],
                    "sourceBlockCode": block_code,
                }
            )
            continue

        if database_backed:
            existing = existing_identities.get(clean_record["blockCode"])
            if existing is not None and strategy == "skip":
                imported_blocks.append({"blockCode": block_code, "name": block_name, "action": "skipped", "row": row_number})
                continue

            if existing is not None:
                normalized["id"] = existing["id"]
                normalized["createdAt"] = existing.get("createdAt", normalized["createdAt"])
                normalized["deletedAt"] = existing.get("deletedAt")
                action = "updated"
            else:
                action = "created"
            stored_block = sanitize_block_for_ledger(normalized)
            pending_database_blocks.append(stored_block)
            existing_identities[block_code] = {
                "id": stored_block.get("id"),
                "blockCode": block_code,
                "createdAt": stored_block.get("createdAt"),
                "deletedAt": stored_block.get("deletedAt"),
            }
            pending_right_blocks.append(normalized)
            imported_blocks.append({"blockCode": block_code, "name": block_name, "action": action, "row": row_number})
            imported_rights_archives.append(
                {
                    "archiveCode": block_code,
                    "linkedBlockCodes": [block_code],
                    "sourceBlockCode": block_code,
                }
            )
            did_change_blocks = True
            continue

        existing_index = active_indexes.get(clean_record["blockCode"])
        if existing_index is not None and strategy == "skip":
            imported_blocks.append({"blockCode": block_code, "name": block_name, "action": "skipped", "row": row_number})
            continue

        if existing_index is not None:
            existing = blocks[existing_index]
            normalized["id"] = existing["id"]
            normalized["createdAt"] = existing.get("createdAt", normalized["createdAt"])
            normalized["deletedAt"] = existing.get("deletedAt")
            blocks[existing_index] = sanitize_block_for_ledger(normalized)
            pending_right_blocks.append(normalized)
            imported_blocks.append({"blockCode": block_code, "name": block_name, "action": "updated", "row": row_number})
            imported_rights_archives.append(
                {
                    "archiveCode": block_code,
                    "linkedBlockCodes": [block_code],
                    "sourceBlockCode": block_code,
                }
            )
            did_change_blocks = True
            continue

        blocks.append(sanitize_block_for_ledger(normalized))
        active_indexes[clean_record["blockCode"]] = len(blocks) - 1
        pending_right_blocks.append(normalized)
        imported_blocks.append({"blockCode": block_code, "name": block_name, "action": "created", "row": row_number})
        imported_rights_archives.append(
            {
                "archiveCode": block_code,
                "linkedBlockCodes": [block_code],
                "sourceBlockCode": block_code,
            }
        )
        did_change_blocks = True

    if did_change_blocks:
        if database_backed:
            save_blocks(pending_database_blocks)
        else:
            save_blocks(blocks)
    upsert_right_archives_from_blocks(pending_right_blocks)

    report = build_report(
        batch_id=batch_id,
        file_name=file_name,
        total_rows=len(records),
        valid_rows=valid_rows,
        invalid_rows=len(records) - valid_rows,
        errors=errors,
        imported_blocks=imported_blocks,
        imported_rights_archives=imported_rights_archives,
    )
    report["createdBy"] = getattr(context, "user", "")
    append_import_batch_audit_event(
        report,
        action="import",
        actor=getattr(context, "user", ""),
        summary={
            "fileName": file_name,
            "strategy": strategy,
            "totalRows": report.get("totalRows"),
            "validRows": report.get("validRows"),
            "invalidRows": report.get("invalidRows"),
        },
        timestamp=report.get("completedAt"),
    )
    save_import_report(report)
    return {**report, "report": report}


@router.get("/imports/forest-blocks/workflow-summary")
def get_forest_block_import_workflow_summary(request: Request) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_VIEW_PERMISSION)
    return import_workflow_summary(context)


@router.get("/imports/forest-blocks/workflow-summary.json")
def export_forest_block_import_workflow_summary(request: Request) -> Response:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION)
    payload = {**import_workflow_summary(context), "exportedAt": now_iso()}
    return Response(
        content=json.dumps(payload, ensure_ascii=False, indent=2),
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="import-workflow-summary.json"'},
    )


@router.get("/imports/forest-blocks/operation-queue")
def get_forest_block_import_operation_queue(
    request: Request,
    limit: int = Query(default=5, ge=1, le=20),
) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_VIEW_PERMISSION)
    return import_operation_queue(limit=limit, context=context)


@router.get("/imports/forest-blocks/delivery-packages")
def list_forest_block_import_delivery_packages(
    request: Request,
    q: str = Query(default=""),
    status: str = Query(default=""),
    acceptanceStatus: str = Query(default=""),
    deliveryStatus: str = Query(default=""),
    linkedBlockCode: str = Query(default=""),
    limit: int = Query(default=100, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_VIEW_PERMISSION)
    return list_import_delivery_packages(
        q=q,
        status=status,
        acceptance_status=acceptanceStatus,
        delivery_status=deliveryStatus,
        linked_block_code=linkedBlockCode,
        limit=limit,
        offset=offset,
        context=context,
    )


@router.get("/imports/forest-blocks/delivery-packages.csv")
def export_forest_block_import_delivery_packages_csv(
    request: Request,
    q: str = Query(default=""),
    status: str = Query(default=""),
    acceptanceStatus: str = Query(default=""),
    deliveryStatus: str = Query(default=""),
    linkedBlockCode: str = Query(default=""),
    limit: int = Query(default=1000, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
) -> Response:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION)
    payload = list_import_delivery_packages(
        q=q,
        status=status,
        acceptance_status=acceptanceStatus,
        delivery_status=deliveryStatus,
        linked_block_code=linkedBlockCode,
        limit=limit,
        offset=offset,
        context=context,
    )
    return csv_download_response(
        "import-delivery-packages.csv",
        [
            "batchId",
            "fileName",
            "status",
            "packageStatus",
            "acceptanceStatus",
            "deliveryStatus",
            "linkedSceneCount",
            "deliveredSceneCount",
            "pendingSceneCount",
            "publishedLayerCount",
            "linkedBlockCount",
            "rightArchiveCount",
            "blockingReasons",
            "updatedAt",
        ],
        payload.get("items") or [],
    )


@router.get("/imports/forest-blocks/delivery-packages.json")
def export_forest_block_import_delivery_packages_json(
    request: Request,
    q: str = Query(default=""),
    status: str = Query(default=""),
    acceptanceStatus: str = Query(default=""),
    deliveryStatus: str = Query(default=""),
    linkedBlockCode: str = Query(default=""),
    limit: int = Query(default=1000, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
) -> Response:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION)
    payload = list_import_delivery_packages(
        q=q,
        status=status,
        acceptance_status=acceptanceStatus,
        delivery_status=deliveryStatus,
        linked_block_code=linkedBlockCode,
        limit=limit,
        offset=offset,
        context=context,
    )
    return json_download_response(
        "import-delivery-packages.json",
        import_delivery_packages_receipt(payload, context),
    )


@router.get("/imports/forest-blocks/sources")
def list_forest_block_import_sources(request: Request) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_VIEW_PERMISSION)
    items = list_import_source_files()
    return {"items": items, "total": len(items)}


@router.post("/imports/forest-blocks/sources/import")
def import_forest_block_source(payload: ImportSourceRequest, request: Request) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_CREATE_PERMISSION)
    source_path = resolve_import_source_path(payload.path)
    records = parse_import_file(source_path.name, source_path.read_bytes())
    result = execute_forest_block_import(
        records=records,
        file_name=source_path.name,
        strategy=payload.strategy,
        context=context,
    )
    return {**result, "source": import_source_record(source_path, next(root for root in allowed_import_source_roots() if path_is_relative_to(source_path, root)))}


@router.post("/imports/forest-blocks")
async def import_forest_blocks(
    request: Request,
    file: UploadFile = File(...),
    strategy: str = Form(default="upsert"),
) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_CREATE_PERMISSION)

    content = await file.read()
    records = parse_import_file(file.filename or "upload", content)
    return execute_forest_block_import(
        records=records,
        file_name=file.filename or "upload",
        strategy=strategy,
        context=context,
    )


@router.get("/imports/forest-blocks/batches")
def list_forest_block_import_batches(
    request: Request,
    q: str = Query(default=""),
    status: str = Query(default=""),
    reviewStatus: str = Query(default=""),
    acceptanceStatus: str = Query(default=""),
    qualityStatus: str = Query(default=""),
    publishRiskStatus: str = Query(default=""),
    sceneId: str = Query(default=""),
    workflowQueue: str = Query(default=""),
    includeDeleted: bool = Query(default=False),
    limit: int = Query(default=100, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    context = request_context(request)
    include_deleted_records = includeDeleted or status == "deleted"
    if include_deleted_records:
        require_permission(context, IMPORT_FOREST_BLOCKS_RESTORE_PERMISSION)
    else:
        require_permission(context, IMPORT_FOREST_BLOCKS_VIEW_PERMISSION)
    return list_import_reports(
        ImportBatchFilters(
            q=q,
            status=status,
            reviewStatus=reviewStatus,
            acceptanceStatus=acceptanceStatus,
            qualityStatus=qualityStatus,
            publishRiskStatus=publishRiskStatus,
            sceneId=sceneId,
            workflowQueue=workflowQueue,
            includeDeleted=include_deleted_records,
            limit=limit,
            offset=offset,
        ),
        context,
    )


@router.get("/imports/forest-blocks/audit-events")
def list_forest_block_import_audit_events(
    request: Request,
    q: str = Query(default=""),
    action: str = Query(default=""),
    batchId: str = Query(default=""),
    includeDeleted: bool = Query(default=False),
    limit: int = Query(default=100, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    context = request_context(request)
    if includeDeleted:
        require_permission(context, IMPORT_FOREST_BLOCKS_RESTORE_PERMISSION)
    else:
        require_permission(context, IMPORT_FOREST_BLOCKS_VIEW_PERMISSION)
    return list_import_audit_events(
        q=q,
        action=action,
        batch_id=batchId,
        include_deleted=includeDeleted,
        limit=limit,
        offset=offset,
        context=context,
    )


@router.get("/imports/forest-blocks/audit-events.csv")
def export_forest_block_import_audit_events_csv(
    request: Request,
    q: str = Query(default=""),
    action: str = Query(default=""),
    batchId: str = Query(default=""),
    includeDeleted: bool = Query(default=False),
    limit: int = Query(default=1000, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
) -> Response:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION)
    if includeDeleted:
        require_permission(context, IMPORT_FOREST_BLOCKS_RESTORE_PERMISSION)
    payload = list_import_audit_events(
        q=q,
        action=action,
        batch_id=batchId,
        include_deleted=includeDeleted,
        limit=limit,
        offset=offset,
        context=context,
    )
    return csv_download_response(
        "import-audit-events.csv",
        ["batchId", "action", "actor", "at", "batchStatus", "acceptanceStatus", "fileName", "summary"],
        payload.get("items") or [],
    )


@router.get("/imports/forest-blocks/quality-issues")
def list_forest_block_import_quality_issues(
    request: Request,
    q: str = Query(default=""),
    issueType: str = Query(default=""),
    severity: str = Query(default=""),
    batchId: str = Query(default=""),
    status: str = Query(default=""),
    includeDeleted: bool = Query(default=False),
    limit: int = Query(default=100, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    context = request_context(request)
    if includeDeleted:
        require_permission(context, IMPORT_FOREST_BLOCKS_RESTORE_PERMISSION)
    else:
        require_permission(context, IMPORT_FOREST_BLOCKS_VIEW_PERMISSION)
    return list_import_quality_issues(
        q=q,
        issue_type=issueType,
        severity=severity,
        batch_id=batchId,
        status=status,
        include_deleted=includeDeleted,
        limit=limit,
        offset=offset,
        context=context,
    )


@router.get("/imports/forest-blocks/quality-issues.csv")
def export_forest_block_import_quality_issues_csv(
    request: Request,
    q: str = Query(default=""),
    issueType: str = Query(default=""),
    severity: str = Query(default=""),
    batchId: str = Query(default=""),
    status: str = Query(default=""),
    includeDeleted: bool = Query(default=False),
    limit: int = Query(default=1000, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
) -> Response:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION)
    if includeDeleted:
        require_permission(context, IMPORT_FOREST_BLOCKS_RESTORE_PERMISSION)
    payload = list_import_quality_issues(
        q=q,
        issue_type=issueType,
        severity=severity,
        batch_id=batchId,
        status=status,
        include_deleted=includeDeleted,
        limit=limit,
        offset=offset,
        context=context,
    )
    return csv_download_response(
        "import-quality-issues.csv",
        [
            "issueId",
            "batchId",
            "fileName",
            "issueType",
            "issueKey",
            "severity",
            "status",
            "sceneId",
            "blockCodes",
            "message",
            "actionRequired",
            "handledBy",
            "handledAt",
            "handlingComment",
            "adminHref",
        ],
        payload.get("items") or [],
    )


@router.patch("/imports/forest-blocks/quality-issues/{issue_id:path}")
def update_forest_block_import_quality_issue(
    issue_id: str,
    payload: ImportQualityIssueUpdateRequest,
    request: Request,
) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_QUALITY_PERMISSION)
    status = payload.status.strip()
    if status not in QUALITY_ISSUE_STATUSES:
        raise HTTPException(status_code=400, detail=f"status must be one of: {', '.join(sorted(QUALITY_ISSUE_STATUSES))}")

    batch_id = issue_id.split(":", 1)[0]
    report = get_report_or_404(batch_id, include_targets=not use_mysql())
    require_import_report_allowed(context, report)
    current_issue = next((issue for issue in quality_issues_for_report(report) if str(issue.get("issueId")) == issue_id), None)
    if not current_issue:
        raise HTTPException(status_code=404, detail="Import quality issue not found")

    event = {
        "at": now_iso(),
        "issueId": issue_id,
        "status": status,
        "comment": payload.comment.strip(),
        "actor": context.user,
    }
    events = list(report.get("qualityIssueEvents") or [])
    events.append(event)
    report["qualityIssueEvents"] = events
    updated_issue = apply_quality_issue_event(report, current_issue)
    append_import_batch_audit_event(
        report,
        action="quality-issue-update",
        actor=context.user,
        summary={
            "issueId": issue_id,
            "status": status,
            "comment": event["comment"],
        },
        timestamp=event["at"],
    )
    save_import_report(report)
    return {
        "ok": True,
        "issue": updated_issue,
        "event": event,
        "report": public_import_report(report),
    }


def get_report_or_404(
    batch_id: str,
    *,
    include_targets: bool = True,
) -> dict[str, Any]:
    database_backed = use_mysql() or use_postgis()
    report = None if database_backed else IMPORT_REPORTS.get(batch_id)
    if report is None:
        if use_mysql():
            report = (
                load_import_report_mysql(batch_id)
                if include_targets
                else load_import_report_mysql(batch_id, include_targets=False)
            )
        elif use_postgis():
            report = load_import_report_postgis(batch_id)
        else:
            report = load_import_report_json(batch_id)
    if report is None:
        raise HTTPException(status_code=404, detail="Import batch not found")
    if not database_backed:
        IMPORT_REPORTS[batch_id] = report
    return report


def mysql_import_batch_visible(context: AuthContext, batch_id: str) -> bool:
    if not context_has_import_batch_scope(context):
        return True
    params: list[Any] = []
    scope_sql = mysql_import_scope_exists_sql(context, params)
    with mysql_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT 1 FROM import_batches WHERE id = %s AND {scope_sql} LIMIT 1",
                tuple([batch_id, *params]),
            )
            return cur.fetchone() is not None


def require_import_batch_summary_read_permission(
    report: dict[str, Any],
    request: Request,
) -> None:
    context = request_context(request)
    if report.get("deletedAt") or report.get("status") == "deleted":
        require_permission(context, IMPORT_FOREST_BLOCKS_RESTORE_PERMISSION)
    else:
        require_permission(context, IMPORT_FOREST_BLOCKS_VIEW_PERMISSION)
    if use_mysql():
        if not mysql_import_batch_visible(context, str(report.get("id") or "")):
            raise HTTPException(status_code=404, detail="Import batch not found")
        return
    require_import_report_allowed(context, report)


def list_import_batch_targets_mysql(
    batch_id: str,
    *,
    kind: str,
    q: str = "",
    limit: int = 100,
    offset: int = 0,
) -> dict[str, Any]:
    query_text = f"%{q.strip()}%"
    if kind == "blocks":
        from_sql = (
            " FROM import_batch_block_links links "
            "JOIN forest_blocks blocks ON blocks.id = links.forest_block_id "
            "WHERE links.import_batch_id = %s"
        )
        params: list[Any] = [batch_id]
        if q.strip():
            from_sql += " AND (blocks.block_code LIKE %s OR blocks.name LIKE %s)"
            params.extend([query_text, query_text])
        select_sql = (
            "SELECT links.target_json, blocks.block_code, blocks.name, links.import_action"
            + from_sql
            + " ORDER BY blocks.block_code LIMIT %s OFFSET %s"
        )
        normalizer = normalize_mysql_import_block_target
    else:
        from_sql = (
            " FROM import_batch_right_links links "
            "JOIN forest_rights rights ON rights.id = links.forest_right_id "
            "WHERE links.import_batch_id = %s"
        )
        params = [batch_id]
        if q.strip():
            from_sql += (
                " AND (rights.archive_code LIKE %s OR rights.certificate_no LIKE %s "
                "OR rights.holder LIKE %s)"
            )
            params.extend([query_text, query_text, query_text])
        select_sql = (
            "SELECT links.target_json, rights.archive_code"
            + from_sql
            + " ORDER BY rights.archive_code LIMIT %s OFFSET %s"
        )
        normalizer = normalize_mysql_import_right_target
    with mysql_connect() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*)" + from_sql, tuple(params))
            total_row = cur.fetchone()
            cur.execute(select_sql, tuple([*params, limit, offset]))
            items = [normalizer(row) for row in cur.fetchall()]
    return {
        "kind": kind,
        "items": items,
        "total": int(total_row[0] if total_row else 0),
        "limit": limit,
        "offset": offset,
    }


def list_import_batch_targets_json(
    report: dict[str, Any],
    *,
    kind: str,
    q: str = "",
    limit: int = 100,
    offset: int = 0,
) -> dict[str, Any]:
    source_key = "importedBlocks" if kind == "blocks" else "importedRightsArchives"
    query = q.strip().lower()
    items = [item for item in report.get(source_key) or [] if isinstance(item, dict)]
    if query:
        items = [
            item
            for item in items
            if query in json.dumps(item, ensure_ascii=False).lower()
        ]
    return {
        "kind": kind,
        "items": items[offset : offset + limit],
        "total": len(items),
        "limit": limit,
        "offset": offset,
    }


def unique_values(values: list[Any]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def import_report_block_codes(report: dict[str, Any]) -> list[str]:
    return unique_values(
        [
            item.get("blockCode")
            for item in report.get("importedBlocks") or []
            if item.get("action") in {"created", "updated"}
        ]
    )


def import_report_right_archive_codes(report: dict[str, Any]) -> list[str]:
    return unique_values([item.get("archiveCode") for item in report.get("importedRightsArchives") or []])


def import_receipt_export_metadata(context: AuthContext | None, permission: str) -> dict[str, Any]:
    return {
        "exportedBy": getattr(context, "user", "") if context else "",
        "exportPermission": permission,
        "exportRoles": role_codes_for_context(context) if context else [],
        "exportDataScopes": effective_data_scopes_for_context(context) if context else {},
    }


def import_batch_acceptance_receipt(report: dict[str, Any], context: AuthContext | None = None) -> dict[str, Any]:
    imagery_links = [dict(link) for link in report.get("imageryLinks") or [] if isinstance(link, dict)]
    quality_issues = quality_issues_for_report(report)
    return {
        "receiptType": "import-batch-acceptance",
        "exportedAt": now_iso(),
        **import_receipt_export_metadata(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION),
        "batch": report,
        "summary": {
            "batchId": report.get("id") or "",
            "fileName": report.get("fileName") or "",
            "status": report.get("status") or "",
            "reviewStatus": report.get("reviewStatus") or "pending",
            "acceptanceStatus": report.get("acceptanceStatus") or "pending",
            "acceptedBy": report.get("acceptedBy") or "",
            "acceptedAt": report.get("acceptedAt"),
            "qualityStatus": report.get("qualityStatus") or "pending",
            "publishRiskStatus": report.get("publishRiskStatus") or "unknown",
            "totalRows": report.get("totalRows") or 0,
            "validRows": report.get("validRows") or 0,
            "invalidRows": report.get("invalidRows") or 0,
            "importedBlockCount": len(import_report_block_codes(report)),
            "rightArchiveCount": len(import_report_right_archive_codes(report)),
            "linkedSceneCount": len(imagery_links),
            "qualityIssueCount": len(quality_issues),
            "auditEventCount": len(report.get("auditEvents") or []),
        },
        "imageryLinks": imagery_links,
        "qualityIssues": quality_issues,
        "auditEvents": list(report.get("auditEvents") or []),
        "reviewEvents": list(report.get("reviewEvents") or []),
        "acceptanceEvents": list(report.get("acceptanceEvents") or []),
    }


def review_import_batch_report(
    report: dict[str, Any],
    payload: ImportBatchReviewRequest,
    request: Request,
) -> tuple[dict[str, Any], dict[str, Any]]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_REVIEW_PERMISSION)
    require_import_report_allowed(context, report)
    if report.get("deletedAt") or report.get("status") == "deleted":
        raise HTTPException(status_code=409, detail="Deleted import batch cannot be reviewed")
    if report.get("status") == "rolled_back":
        raise HTTPException(status_code=409, detail="Rolled back import batch cannot be reviewed")

    decision = payload.decision.strip()
    allowed_decisions = {"approved", "rejected", "needs_correction"}
    if decision not in allowed_decisions:
        raise HTTPException(status_code=400, detail="decision must be approved, rejected, or needs_correction")

    timestamp = now_iso()
    event = {
        "at": timestamp,
        "action": "review",
        "decision": decision,
        "actor": getattr(context, "user", ""),
        "comment": payload.comment.strip(),
        "previousReviewStatus": report.get("reviewStatus") or "pending",
    }
    events = list(report.get("reviewEvents") or [])
    events.append(event)
    report["reviewStatus"] = decision
    report["reviewComment"] = payload.comment.strip()
    report["reviewedAt"] = timestamp
    report["reviewedBy"] = getattr(context, "user", "")
    report["reviewEvents"] = events
    append_import_batch_audit_event(
        report,
        action="review",
        actor=getattr(context, "user", ""),
        summary={
            "reviewStatus": decision,
            "comment": payload.comment.strip(),
            "previousReviewStatus": event["previousReviewStatus"],
        },
        timestamp=timestamp,
    )
    save_import_report(report)
    return report, event


def update_import_batch_acceptance(
    report: dict[str, Any],
    payload: ImportBatchAcceptanceRequest,
    request: Request,
) -> tuple[dict[str, Any], dict[str, Any]]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_ACCEPTANCE_PERMISSION)
    require_import_report_allowed(context, report)
    if report.get("deletedAt") or report.get("status") == "deleted":
        raise HTTPException(status_code=409, detail="Deleted import batch cannot be accepted")
    if report.get("status") == "rolled_back":
        raise HTTPException(status_code=409, detail="Rolled back import batch cannot be accepted")

    status = payload.status.strip()
    if status not in IMPORT_ACCEPTANCE_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"status must be one of: {', '.join(sorted(IMPORT_ACCEPTANCE_STATUSES))}",
        )
    if status != "pending" and str(report.get("reviewStatus") or "pending") != "approved":
        raise HTTPException(status_code=409, detail="Import batch review approved status is required before acceptance")

    timestamp = now_iso()
    actor = getattr(context, "user", "")
    previous_status = report.get("acceptanceStatus") or "pending"
    event = {
        "at": timestamp,
        "action": "acceptance",
        "status": status,
        "actor": actor,
        "comment": payload.comment.strip(),
        "previousAcceptanceStatus": previous_status,
    }
    events = list(report.get("acceptanceEvents") or [])
    events.append(event)
    report["acceptanceStatus"] = status
    report["acceptanceComment"] = payload.comment.strip()
    report["acceptedAt"] = timestamp
    report["acceptedBy"] = actor
    report["acceptanceEvents"] = events
    append_import_batch_audit_event(
        report,
        action="acceptance",
        actor=actor,
        summary={
            "acceptanceStatus": status,
            "comment": payload.comment.strip(),
            "previousAcceptanceStatus": previous_status,
        },
        timestamp=timestamp,
    )
    save_import_report(report)
    return report, event


def parse_scene_bounds(value: Any) -> list[float] | None:
    if isinstance(value, str):
        raw_parts: list[Any] = [part.strip() for part in value.split(",")]
    elif isinstance(value, (list, tuple)):
        raw_parts = list(value)
    else:
        return None
    if len(raw_parts) != 4:
        return None
    try:
        west, south, east, north = [float(part) for part in raw_parts]
    except (TypeError, ValueError):
        return None
    if west >= east or south >= north:
        return None
    return [west, south, east, north]


def geometry_coordinate_points(geometry: dict[str, Any] | None) -> list[tuple[float, float]]:
    if not isinstance(geometry, dict):
        return []
    geometry_type = str(geometry.get("type") or "")
    coordinates = geometry.get("coordinates") or []
    points: list[tuple[float, float]] = []

    def add_point(point: Any) -> None:
        if not isinstance(point, (list, tuple)) or len(point) < 2:
            return
        try:
            points.append((float(point[0]), float(point[1])))
        except (TypeError, ValueError):
            return

    if geometry_type == "Polygon":
        for ring in coordinates:
            for point in ring or []:
                add_point(point)
    elif geometry_type == "MultiPolygon":
        for polygon in coordinates:
            for ring in polygon or []:
                for point in ring or []:
                    add_point(point)
    return points


def geometry_bounds(geometry: dict[str, Any] | None) -> list[float] | None:
    points = geometry_coordinate_points(geometry)
    if not points:
        return None
    xs = [point[0] for point in points]
    ys = [point[1] for point in points]
    return [min(xs), min(ys), max(xs), max(ys)]


def bounds_intersect(left: list[float], right: list[float]) -> bool:
    left_west, left_south, left_east, left_north = left
    right_west, right_south, right_east, right_north = right
    return left_east >= right_west and left_west <= right_east and left_north >= right_south and left_south <= right_north


def import_batch_coverage_check(scene: dict[str, Any], linked_block_records: list[dict[str, Any]]) -> dict[str, Any]:
    scene_bounds = parse_scene_bounds(scene.get("bounds"))
    missing_geometry_codes: list[str] = []
    outside_scene_bounds_codes: list[str] = []
    checked_blocks = 0
    for block in linked_block_records:
        block_code = str(block.get("blockCode") or block.get("code") or block.get("id") or "").strip()
        block_bounds = geometry_bounds(block.get("geometry"))
        if block_bounds is None:
            missing_geometry_codes.append(block_code)
            continue
        if scene_bounds is None:
            continue
        checked_blocks += 1
        if not bounds_intersect(block_bounds, scene_bounds):
            outside_scene_bounds_codes.append(block_code)

    warnings: list[str] = []
    if scene_bounds is None:
        warnings.append("missing_scene_bounds")
    if missing_geometry_codes:
        warnings.append("missing_geometry")
    if outside_scene_bounds_codes:
        warnings.append("outside_scene_bounds")

    return {
        "status": "warning" if warnings else "pass",
        "sceneHasBounds": scene_bounds is not None,
        "sceneBounds": scene_bounds or [],
        "totalBlocks": len(linked_block_records),
        "checkedBlocks": checked_blocks,
        "missingGeometryCount": len(missing_geometry_codes),
        "outsideSceneBoundsCount": len(outside_scene_bounds_codes),
        "missingGeometryBlockCodes": missing_geometry_codes,
        "outsideSceneBoundsBlockCodes": outside_scene_bounds_codes,
        "warnings": warnings,
    }


def mysql_import_batch_scene_plan(
    batch_id: str,
    scene: dict[str, Any],
    *,
    sample_limit: int = IMPORT_RELATION_EVENT_SAMPLE_LIMIT,
    connection_factory: Any = mysql_connect,
) -> dict[str, Any]:
    """Build a bounded-memory publish plan from normalized import relations."""
    scene_bounds = parse_scene_bounds(scene.get("bounds"))
    outside_sql = "0"
    aggregate_params: list[Any] = []
    if scene_bounds is not None:
        west, south, east, north = scene_bounds
        outside_sql = (
            "SUM(CASE WHEN blocks.deleted_at IS NULL "
            "AND geometries.forest_block_id IS NOT NULL "
            "AND (geometries.max_longitude < %s OR geometries.min_longitude > %s "
            "OR geometries.max_latitude < %s OR geometries.min_latitude > %s) "
            "THEN 1 ELSE 0 END)"
        )
        aggregate_params.extend([west, east, south, north])
    aggregate_params.append(batch_id)
    base_from = (
        " FROM import_batch_block_links links "
        "JOIN forest_blocks blocks ON blocks.id = links.forest_block_id "
        "LEFT JOIN forest_block_geometries geometries ON geometries.forest_block_id = blocks.id "
        "WHERE links.import_batch_id = %s "
        "AND links.import_action IN ('created', 'updated')"
    )
    with connection_factory() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*), "
                "SUM(CASE WHEN blocks.deleted_at IS NULL THEN 1 ELSE 0 END), "
                "SUM(CASE WHEN blocks.deleted_at IS NOT NULL THEN 1 ELSE 0 END), "
                "SUM(CASE WHEN blocks.deleted_at IS NULL AND geometries.forest_block_id IS NULL THEN 1 ELSE 0 END), "
                f"{outside_sql}{base_from}",
                tuple(aggregate_params),
            )
            aggregate = cur.fetchone() or (0, 0, 0, 0, 0)
            total_count = int(aggregate[0] or 0)
            active_count = int(aggregate[1] or 0)
            missing_count = int(aggregate[2] or 0)
            missing_geometry_count = int(aggregate[3] or 0)
            outside_count = int(aggregate[4] or 0)

            cur.execute(
                "SELECT blocks.id, blocks.block_code "
                + base_from
                + " AND blocks.deleted_at IS NULL ORDER BY blocks.block_code LIMIT %s",
                (batch_id, sample_limit),
            )
            linked_blocks = [
                {"blockId": str(row[0]), "blockCode": str(row[1] or "")}
                for row in cur.fetchall()
            ]
            cur.execute(
                "SELECT blocks.block_code "
                + base_from
                + " AND blocks.deleted_at IS NOT NULL ORDER BY blocks.block_code LIMIT %s",
                (batch_id, sample_limit),
            )
            skipped_blocks = [
                {"blockCode": str(row[0] or ""), "reason": "block_not_found"}
                for row in cur.fetchall()
            ]
            missing_geometry_codes: list[str] = []
            if missing_geometry_count:
                cur.execute(
                    "SELECT blocks.block_code "
                    + base_from
                    + " AND blocks.deleted_at IS NULL AND geometries.forest_block_id IS NULL "
                    "ORDER BY blocks.block_code LIMIT %s",
                    (batch_id, sample_limit),
                )
                missing_geometry_codes = [str(row[0] or "") for row in cur.fetchall()]
            outside_codes: list[str] = []
            if outside_count and scene_bounds is not None:
                west, south, east, north = scene_bounds
                cur.execute(
                    "SELECT blocks.block_code "
                    + base_from
                    + " AND blocks.deleted_at IS NULL AND geometries.forest_block_id IS NOT NULL "
                    "AND (geometries.max_longitude < %s OR geometries.min_longitude > %s "
                    "OR geometries.max_latitude < %s OR geometries.min_latitude > %s) "
                    "ORDER BY blocks.block_code LIMIT %s",
                    (batch_id, west, east, south, north, sample_limit),
                )
                outside_codes = [str(row[0] or "") for row in cur.fetchall()]
            cur.execute(
                "SELECT COUNT(*) FROM import_batch_right_links links "
                "JOIN forest_rights rights ON rights.id = links.forest_right_id "
                "WHERE links.import_batch_id = %s AND rights.deleted_at IS NULL",
                (batch_id,),
            )
            right_count_row = cur.fetchone()
            right_count = int((right_count_row or (0,))[0] or 0)
            cur.execute(
                "SELECT rights.archive_code FROM import_batch_right_links links "
                "JOIN forest_rights rights ON rights.id = links.forest_right_id "
                "WHERE links.import_batch_id = %s AND rights.deleted_at IS NULL "
                "ORDER BY rights.archive_code LIMIT %s",
                (batch_id, sample_limit),
            )
            right_codes = [str(row[0] or "") for row in cur.fetchall()]

    warnings: list[str] = []
    if scene_bounds is None:
        warnings.append("missing_scene_bounds")
    if missing_geometry_count:
        warnings.append("missing_geometry")
    if outside_count:
        warnings.append("outside_scene_bounds")
    checked_blocks = active_count - missing_geometry_count if scene_bounds is not None else 0
    coverage_check = {
        "status": "warning" if warnings else "pass",
        "sceneHasBounds": scene_bounds is not None,
        "sceneBounds": scene_bounds or [],
        "totalBlocks": active_count,
        "checkedBlocks": max(0, checked_blocks),
        "missingGeometryCount": missing_geometry_count,
        "outsideSceneBoundsCount": outside_count,
        "missingGeometryBlockCodes": missing_geometry_codes,
        "outsideSceneBoundsBlockCodes": outside_codes,
        "targetsTruncated": active_count > sample_limit,
        "warnings": warnings,
    }
    return {
        "targetCount": total_count,
        "linkedBlockCount": active_count,
        "missingBlockCount": missing_count,
        "linkedBlocks": linked_blocks,
        "skippedBlocks": skipped_blocks,
        "linkedRightArchiveCount": right_count,
        "linkedRightArchiveCodes": right_codes,
        "targetsTruncated": active_count > sample_limit or right_count > sample_limit,
        "coverageCheck": coverage_check,
    }


def import_batch_quality_from_coverage(coverage_check: dict[str, Any]) -> dict[str, Any]:
    findings = [str(item) for item in coverage_check.get("warnings") or [] if str(item)]
    quality_status = "warning" if str(coverage_check.get("status") or "") == "warning" else "pass"
    needs_correction = quality_status == "warning"
    return {
        "qualityStatus": quality_status,
        "qualityFindings": findings,
        "reviewRecommendation": "needs_correction" if needs_correction else "approved",
        "publishRiskStatus": "warning" if needs_correction else "clear",
    }


def upsert_batch_scene_link(item: dict[str, Any]) -> None:
    if use_mysql():
        save_scene_link_mysql(item)
        return
    if use_postgis():
        upsert_scene_link_postgis(item)
        return
    records = [
        record
        for record in load_scene_links()
        if not (
            record.get("forestBlockId") == item["forestBlockId"]
            and record.get("sceneId") == item["sceneId"]
            and record.get("relationType") == item["relationType"]
        )
    ]
    records.append(item)
    save_scene_links(records)


def upsert_import_batch_layer(
    *,
    request: Request,
    report: dict[str, Any],
    scene: dict[str, Any],
    payload: ImportBatchSceneLayerRequest,
    context: AuthContext,
    linked_block_codes: list[str],
    linked_right_archive_codes: list[str],
    linked_block_count: int | None = None,
    linked_right_archive_count: int | None = None,
    targets_truncated: bool = False,
    coverage_check: dict[str, Any] | None = None,
) -> dict[str, Any]:
    batch_id = str(report.get("id") or "")
    scene_id = str(scene.get("id") or payload.sceneId).strip()
    layer_id = f"import-layer-{batch_id}"
    record_code = f"IMPORT-LAYER-{batch_id}"
    quality = import_batch_quality_from_coverage(coverage_check or {})
    records = [] if use_mysql() else layer_records()
    existing_index: int | None = None
    existing: dict[str, Any] = find_layer_record_for_upsert(layer_id, record_code) or {}
    if not use_mysql():
        for index, record in enumerate(records):
            if str(record.get("id")) == layer_id or str(record.get("recordCode")) == record_code:
                existing_index = index
                existing = record
                break

    base_url = str(request.base_url).rstrip("/")
    properties = dict(existing.get("properties") or {})
    properties.update(payload.properties or {})
    properties.update(
        {
            "source": "import-batch",
            "importBatchId": batch_id,
            "importFileName": report.get("fileName") or "",
            "sourceSceneId": scene_id,
            "sceneName": scene.get("name") or scene_id,
            "bounds": scene.get("bounds") or [],
            "capturedAt": scene.get("capturedAt") or "",
            "tileUrl": f"{base_url}/api/scenes/{scene_id}/tiles/{{z}}/{{x}}/{{y}}.png",
            "tileJsonUrl": f"{base_url}/api/scenes/{scene_id}/tilejson.json",
            "linkedBlockCount": linked_block_count if linked_block_count is not None else len(linked_block_codes),
            "linkedRightArchiveCount": (
                linked_right_archive_count
                if linked_right_archive_count is not None
                else len(linked_right_archive_codes)
            ),
            "linkedTargetsTruncated": targets_truncated,
            "coverageCheck": coverage_check or {},
            **quality,
        }
    )
    layer_payload = {
        **existing,
        "id": layer_id,
        "recordCode": record_code,
        "name": payload.layerName or f"{report.get('fileName') or batch_id} 影像图层",
        "status": "published",
        "layerType": "imagery",
        "dataSource": f"scene:{scene_id}",
        "style": payload.style or existing.get("style") or {"type": "raster", "opacity": 0.9},
        "zIndex": payload.zIndex if payload.zIndex is not None else existing.get("zIndex"),
        "visibleOnDashboard": True,
        "linkedBlockCodes": linked_block_codes,
        "linkedRightArchiveCodes": linked_right_archive_codes,
        "properties": properties,
        "deletedAt": None,
    }
    normalized = normalize_record(layer_payload, default_status="published")
    require_map_layer_upsert_permissions(context, existing=existing, next_layer=normalized)
    normalized = append_map_layer_audit_event(
        normalized,
        "publish-from-import",
        context,
        before=existing or None,
        changed_fields=map_layer_changed_fields(existing, normalized),
    )
    if use_mysql():
        upsert_import_batch_layer_mysql(normalized, import_batch_id=batch_id)
    elif existing_index is None:
        records.append(normalized)
        save_layer_records(records)
    else:
        records[existing_index] = normalized
        save_layer_records(records)
    return enrich_map_layer_record(normalized)


def link_import_batch_scene_layer_records(
    report: dict[str, Any],
    payload: ImportBatchSceneLayerRequest,
    request: Request,
) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_SCENE_LAYER_LINK_PERMISSION)
    require_import_report_allowed(context, report)
    if report.get("deletedAt") or report.get("status") == "deleted":
        raise HTTPException(status_code=409, detail="Deleted import batch cannot link imagery")
    if report.get("status") == "rolled_back":
        raise HTTPException(status_code=409, detail="Rolled back import batch cannot link imagery")
    if str(report.get("reviewStatus") or "pending") != "approved":
        raise HTTPException(status_code=409, detail="Import batch review approved status is required before linking imagery")

    scene = require_visible_scene(payload.sceneId.strip(), context)
    relation_type = payload.relationType.strip() or "coverage"
    link_items: list[dict[str, Any]] = []
    if use_mysql():
        mysql_plan = mysql_import_batch_scene_plan(str(report.get("id") or ""), scene)
        if not mysql_plan["targetCount"]:
            raise HTTPException(status_code=409, detail="Import batch has no forest blocks to link")
        if mysql_plan["missingBlockCount"]:
            raise HTTPException(status_code=409, detail="Import batch has missing imported forest blocks")
        linked_block_count = int(mysql_plan["linkedBlockCount"])
        if not linked_block_count:
            raise HTTPException(status_code=409, detail="No active forest blocks were linked")
        linked_blocks = [
            {
                **item,
                "sceneId": scene["id"],
                "relationType": relation_type,
            }
            for item in mysql_plan["linkedBlocks"]
        ]
        skipped_blocks = list(mysql_plan["skippedBlocks"])
        linked_block_codes = [item["blockCode"] for item in linked_blocks]
        linked_right_archive_codes = list(mysql_plan["linkedRightArchiveCodes"])
        linked_right_archive_count = int(mysql_plan["linkedRightArchiveCount"])
        targets_truncated = bool(mysql_plan["targetsTruncated"])
        coverage_check = dict(mysql_plan["coverageCheck"])
    else:
        block_codes = import_report_block_codes(report)
        if not block_codes:
            raise HTTPException(status_code=409, detail="Import batch has no forest blocks to link")
        linked_blocks = []
        linked_block_records: list[dict[str, Any]] = []
        skipped_blocks = []
        for block_code in block_codes:
            block = block_by_code(block_code)
            if not block:
                skipped_blocks.append({"blockCode": block_code, "reason": "block_not_found"})
                continue
            link_item = {
                "forestBlockId": block["id"],
                "sceneId": scene["id"],
                "relationType": relation_type,
                "capturedAt": payload.capturedAt or scene.get("capturedAt") or None,
                "confidence": payload.confidence,
            }
            link_items.append(link_item)
            linked_blocks.append(
                {
                    "blockCode": block_code,
                    "blockId": block["id"],
                    "sceneId": scene["id"],
                    "relationType": relation_type,
                }
            )
            linked_block_records.append(block)
        if skipped_blocks:
            raise HTTPException(status_code=409, detail="Import batch has missing imported forest blocks")
        if not linked_blocks:
            raise HTTPException(status_code=409, detail="No active forest blocks were linked")
        linked_block_codes = [item["blockCode"] for item in linked_blocks]
        linked_block_count = len(linked_block_codes)
        linked_right_archive_codes = import_report_right_archive_codes(report)
        linked_right_archive_count = len(linked_right_archive_codes)
        targets_truncated = False
        coverage_check = import_batch_coverage_check(scene, linked_block_records)
    quality = import_batch_quality_from_coverage(coverage_check)
    if payload.publishLayer:
        require_import_batch_layer_preflight_permissions(context, str(report.get("id") or ""))

    if use_mysql():
        save_import_batch_scene_links_mysql(
            str(report.get("id") or ""),
            scene_id=str(scene["id"]),
            relation_type=relation_type,
            captured_at=payload.capturedAt or scene.get("capturedAt") or None,
            confidence=payload.confidence,
        )
    else:
        for link_item in link_items:
            upsert_batch_scene_link(link_item)

    layer = None
    if payload.publishLayer:
        layer = upsert_import_batch_layer(
            request=request,
            report=report,
            scene=scene,
            payload=payload,
            context=context,
            linked_block_codes=linked_block_codes,
            linked_right_archive_codes=linked_right_archive_codes,
            linked_block_count=linked_block_count,
            linked_right_archive_count=linked_right_archive_count,
            targets_truncated=targets_truncated,
            coverage_check=coverage_check,
        )

    event = {
        "at": now_iso(),
        "actor": getattr(context, "user", ""),
        "sceneId": scene["id"],
        "relationType": relation_type,
        "linkedBlockCodes": linked_block_codes,
        "linkedBlockCount": linked_block_count,
        "linkedRightArchiveCodes": linked_right_archive_codes,
        "linkedRightArchiveCount": linked_right_archive_count,
        "linkedTargetsTruncated": targets_truncated,
        "linkedBlocks": linked_blocks,
        "skippedBlocks": skipped_blocks,
        "coverageCheck": coverage_check,
        "quality": quality,
        "layerId": layer.get("id") if layer else None,
        "layerRecordCode": layer.get("recordCode") if layer else None,
    }
    imagery_links = list(report.get("imageryLinks") or [])
    imagery_links.append(event)
    report["imageryLinks"] = imagery_links
    report.update(quality)
    append_import_batch_audit_event(
        report,
        action="link-scene-layer",
        actor=getattr(context, "user", ""),
        summary={
            "sceneId": scene["id"],
            "relationType": relation_type,
            "linkedBlockCount": linked_block_count,
            "layerRecordCode": event.get("layerRecordCode"),
            "coverageStatus": coverage_check.get("status"),
            **quality,
        },
        timestamp=event["at"],
    )
    save_import_report(report)
    return {
        "ok": True,
        "id": report.get("id"),
        "sceneId": scene["id"],
        "linkedBlocks": linked_blocks,
        "linkedBlockCount": linked_block_count,
        "linkedTargetsTruncated": targets_truncated,
        "skippedBlocks": skipped_blocks,
        "coverageCheck": coverage_check,
        "quality": quality,
        "layer": layer,
        "report": report,
        "event": event,
    }


def readiness_check(key: str, status: str, message: str) -> dict[str, str]:
    return {"key": key, "status": status, "message": message}


def import_batch_publish_readiness(
    report: dict[str, Any],
    payload: ImportBatchSceneLayerRequest,
    request: Request,
) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_SCENE_LAYER_LINK_PERMISSION)
    require_import_report_allowed(context, report)

    batch_id = str(report.get("id") or "")
    scene_id = payload.sceneId.strip()
    checks: list[dict[str, str]] = []
    blocking_reasons: list[str] = []
    warnings: list[str] = []

    def add_check(key: str, ok: bool, message: str, *, warning: bool = False, reason: str | None = None) -> None:
        status = "pass" if ok else ("warning" if warning else "blocked")
        checks.append(readiness_check(key, status, message))
        if not ok and warning:
            warnings.append(reason or key)
        elif not ok:
            blocking_reasons.append(reason or key)

    batch_deleted = bool(report.get("deletedAt")) or report.get("status") == "deleted"
    batch_rolled_back = report.get("status") == "rolled_back"
    add_check("batch_active", not batch_deleted and not batch_rolled_back, str(report.get("status") or "unknown"), reason="batch_not_active")
    add_check("review_approved", str(report.get("reviewStatus") or "pending") == "approved", str(report.get("reviewStatus") or "pending"), reason="review_not_approved")
    add_check("valid_rows", int(report.get("validRows") or 0) > 0, str(report.get("validRows") or 0), reason="no_valid_rows")
    add_check("scene_selected", bool(scene_id), scene_id or "missing", reason="scene_required")

    scene: dict[str, Any] | None = None
    if scene_id:
        try:
            scene = require_visible_scene(scene_id, context)
            add_check("scene_visible", True, scene_id)
        except HTTPException:
            add_check("scene_visible", False, scene_id, reason="scene_not_visible")

    mysql_plan: dict[str, Any] | None = None
    linked_block_records: list[dict[str, Any]] = []
    if use_mysql():
        mysql_plan = mysql_import_batch_scene_plan(batch_id, scene or {"bounds": []})
        target_count = int(mysql_plan["targetCount"])
        linked_block_count = int(mysql_plan["linkedBlockCount"])
        missing_block_count = int(mysql_plan["missingBlockCount"])
        skipped_blocks = list(mysql_plan["skippedBlocks"])
        add_check("block_codes", target_count > 0, str(target_count), reason="no_imported_blocks")
        add_check("active_blocks", linked_block_count > 0, str(linked_block_count), reason="no_active_blocks")
        if missing_block_count:
            add_check("all_imported_blocks_active", False, str(missing_block_count), reason="missing_imported_blocks")
        elif target_count:
            add_check("all_imported_blocks_active", True, "0")
    else:
        block_codes = import_report_block_codes(report)
        add_check("block_codes", bool(block_codes), str(len(block_codes)), reason="no_imported_blocks")
        skipped_blocks = []
        for block_code in block_codes:
            block = block_by_code(block_code)
            if block:
                linked_block_records.append(block)
            else:
                skipped_blocks.append({"blockCode": block_code, "reason": "block_not_found"})
        linked_block_count = len(linked_block_records)
        missing_block_count = len(skipped_blocks)
        add_check("active_blocks", bool(linked_block_records), str(linked_block_count), reason="no_active_blocks")
        if skipped_blocks:
            add_check("all_imported_blocks_active", False, str(missing_block_count), reason="missing_imported_blocks")
        elif block_codes:
            add_check("all_imported_blocks_active", True, "0")

    coverage_check: dict[str, Any] = {}
    quality = {
        "qualityStatus": report.get("qualityStatus") or "pending",
        "qualityFindings": report.get("qualityFindings") or [],
        "reviewRecommendation": report.get("reviewRecommendation") or "",
        "publishRiskStatus": report.get("publishRiskStatus") or "unknown",
    }
    if scene and (linked_block_records or (mysql_plan and linked_block_count)):
        coverage_check = (
            dict(mysql_plan["coverageCheck"])
            if mysql_plan is not None
            else import_batch_coverage_check(scene, linked_block_records)
        )
        quality = import_batch_quality_from_coverage(coverage_check)
        if coverage_check.get("status") == "warning":
            checks.append(readiness_check("coverage_check", "warning", ",".join(coverage_check.get("warnings") or [])))
            warnings.extend(str(item) for item in coverage_check.get("warnings") or [])
        else:
            checks.append(readiness_check("coverage_check", "pass", "pass"))

    ready = not blocking_reasons
    return {
        "ok": True,
        "id": batch_id,
        "sceneId": scene_id,
        "ready": ready,
        "checks": checks,
        "blockingReasons": blocking_reasons,
        "warnings": unique_values(warnings),
        "linkedBlockCount": linked_block_count,
        "missingBlockCount": missing_block_count,
        "skippedBlocks": skipped_blocks,
        "coverageCheck": coverage_check,
        "quality": quality,
        "publishLayer": payload.publishLayer,
    }


def require_import_batch_read_permission(report: dict[str, Any], request: Request) -> None:
    context = request_context(request)
    if report.get("deletedAt") or report.get("status") == "deleted":
        require_permission(context, IMPORT_FOREST_BLOCKS_RESTORE_PERMISSION)
        require_import_report_allowed(context, report)
        return
    require_permission(context, IMPORT_FOREST_BLOCKS_VIEW_PERMISSION)
    require_import_report_allowed(context, report)


def finalize_import_batch_rollback(
    report: dict[str, Any],
    *,
    timestamp: str,
    actor: str,
    rolled_back_blocks: list[dict[str, Any]],
    skipped_blocks: list[dict[str, Any]],
    updated_count: int,
    skipped_count: int,
) -> dict[str, Any]:
    report["status"] = "rolled_back"
    report["rolledBackAt"] = timestamp
    report["updatedAt"] = timestamp
    report["rolledBackBlocks"] = rolled_back_blocks
    report["rollbackSkippedBlocks"] = skipped_blocks
    report["rollbackSummary"] = {
        "blocksSoftDeleted": len(rolled_back_blocks),
        "blocksSkipped": len(skipped_blocks),
        "updatedRowsRequireManualReview": updated_count,
        "skippedRowsIgnored": skipped_count,
    }
    append_import_batch_audit_event(
        report,
        action="rollback",
        actor=actor,
        summary=report["rollbackSummary"],
        timestamp=timestamp,
    )
    save_import_report(report)
    return report


def rollback_import_batch_records_mysql(
    report: dict[str, Any],
    actor: str = "",
    *,
    batch_size: int = 500,
) -> dict[str, Any]:
    batch_id = str(report.get("id") or "")
    timestamp = now_iso()
    rolled_back_blocks: list[dict[str, Any]] = []
    skipped_blocks: list[dict[str, Any]] = []
    eligible_ids: list[str] = []
    with mysql_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT blocks.id, blocks.block_code, blocks.source_batch_id, blocks.deleted_at "
                "FROM import_batch_block_links links "
                "JOIN forest_blocks blocks ON blocks.id = links.forest_block_id "
                "WHERE links.import_batch_id = %s AND links.import_action = 'created' "
                "ORDER BY blocks.block_code FOR UPDATE",
                (batch_id,),
            )
            for row in cur.fetchall():
                if hasattr(row, "keys"):
                    source = dict(row)
                    block_id = str(source.get("id") or "")
                    block_code = str(source.get("block_code") or "")
                    source_batch_id = str(source.get("source_batch_id") or "")
                    deleted_at = source.get("deleted_at")
                else:
                    block_id, block_code, source_batch_id, deleted_at = row
                    block_id = str(block_id or "")
                    block_code = str(block_code or "")
                    source_batch_id = str(source_batch_id or "")
                if deleted_at:
                    skipped_blocks.append({"blockCode": block_code, "reason": "already_deleted"})
                elif source_batch_id != batch_id:
                    skipped_blocks.append(
                        {"blockCode": block_code, "reason": "source_batch_mismatch"}
                    )
                else:
                    eligible_ids.append(block_id)
                    rolled_back_blocks.append(
                        {"blockCode": block_code, "action": "soft_deleted"}
                    )
            cur.execute(
                "SELECT "
                "COALESCE(SUM(import_action = 'updated'), 0), "
                "COALESCE(SUM(import_action = 'skipped'), 0) "
                "FROM import_batch_block_links WHERE import_batch_id = %s",
                (batch_id,),
            )
            count_row = cur.fetchone() or (0, 0)
            updated_count = int(count_row[0] or 0)
            skipped_count = int(count_row[1] or 0)
            batch_size = max(1, int(batch_size))
            for start in range(0, len(eligible_ids), batch_size):
                id_batch = eligible_ids[start : start + batch_size]
                placeholders = ", ".join(["%s"] * len(id_batch))
                cur.execute(
                    "UPDATE forest_blocks SET deleted_at = %s, updated_at = %s "
                    f"WHERE id IN ({placeholders}) AND deleted_at IS NULL",
                    tuple([mysql_datetime(timestamp), mysql_datetime(timestamp), *id_batch]),
                )
        conn.commit()
    return finalize_import_batch_rollback(
        report,
        timestamp=timestamp,
        actor=actor,
        rolled_back_blocks=rolled_back_blocks,
        skipped_blocks=skipped_blocks,
        updated_count=updated_count,
        skipped_count=skipped_count,
    )


def rollback_import_batch_records(report: dict[str, Any], actor: str = "") -> dict[str, Any]:
    if use_mysql():
        return rollback_import_batch_records_mysql(report, actor)
    batch_id = str(report.get("id") or "")
    timestamp = now_iso()
    imported_blocks = report.get("importedBlocks") or []
    created_block_codes = {
        str(item.get("blockCode") or "").strip()
        for item in imported_blocks
        if item.get("action") == "created" and str(item.get("blockCode") or "").strip()
    }
    rolled_back_blocks: list[dict[str, Any]] = []
    skipped_blocks: list[dict[str, Any]] = []
    did_change_blocks = False

    if created_block_codes:
        blocks = load_all_blocks()
        for block in blocks:
            block_code = str(block.get("blockCode") or "").strip()
            if block_code not in created_block_codes:
                continue
            if block.get("deletedAt"):
                skipped_blocks.append({"blockCode": block_code, "reason": "already_deleted"})
                continue
            if str(block.get("sourceBatchId") or "") != batch_id:
                skipped_blocks.append({"blockCode": block_code, "reason": "source_batch_mismatch"})
                continue
            block["deletedAt"] = timestamp
            block["updatedAt"] = timestamp
            rolled_back_blocks.append({"blockCode": block_code, "action": "soft_deleted"})
            did_change_blocks = True
        if did_change_blocks:
            save_blocks(blocks)

    updated_count = len([item for item in imported_blocks if item.get("action") == "updated"])
    skipped_count = len([item for item in imported_blocks if item.get("action") == "skipped"])
    return finalize_import_batch_rollback(
        report,
        timestamp=timestamp,
        actor=actor,
        rolled_back_blocks=rolled_back_blocks,
        skipped_blocks=skipped_blocks,
        updated_count=updated_count,
        skipped_count=skipped_count,
    )


@router.post("/imports/{batch_id}/rollback")
def rollback_import_batch(batch_id: str, request: Request) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_ROLLBACK_PERMISSION)
    report = get_report_or_404(batch_id, include_targets=not use_mysql())
    require_import_report_allowed(context, report)
    if report.get("deletedAt"):
        raise HTTPException(status_code=409, detail="Deleted import batch cannot be rolled back")
    rolled_back = rollback_import_batch_records(report, actor=getattr(context, "user", ""))
    return {
        "ok": True,
        "id": batch_id,
        "status": rolled_back.get("status"),
        "rolledBackAt": rolled_back.get("rolledBackAt"),
        "rolledBackBlocks": rolled_back.get("rolledBackBlocks") or [],
        "rollbackSkippedBlocks": rolled_back.get("rollbackSkippedBlocks") or [],
        "rollbackSummary": rolled_back.get("rollbackSummary") or {},
    }


@router.post("/imports/{batch_id}/review")
def review_import_batch(
    batch_id: str,
    payload: ImportBatchReviewRequest,
    request: Request,
) -> dict[str, Any]:
    report = get_report_or_404(batch_id, include_targets=not use_mysql())
    reviewed_report, event = review_import_batch_report(report, payload, request)
    return {
        "ok": True,
        "id": batch_id,
        "reviewStatus": reviewed_report.get("reviewStatus"),
        "reviewedAt": reviewed_report.get("reviewedAt"),
        "reviewedBy": reviewed_report.get("reviewedBy"),
        "report": public_import_report(reviewed_report),
        "event": event,
    }


@router.post("/imports/{batch_id}/acceptance")
def accept_import_batch(
    batch_id: str,
    payload: ImportBatchAcceptanceRequest,
    request: Request,
) -> dict[str, Any]:
    report = get_report_or_404(batch_id, include_targets=not use_mysql())
    accepted_report, event = update_import_batch_acceptance(report, payload, request)
    return {
        "ok": True,
        "id": batch_id,
        "acceptanceStatus": accepted_report.get("acceptanceStatus"),
        "acceptedAt": accepted_report.get("acceptedAt"),
        "acceptedBy": accepted_report.get("acceptedBy"),
        "report": public_import_report(accepted_report),
        "event": event,
    }


@router.post("/imports/{batch_id}/link-scene-layer")
def link_import_batch_scene_layer(
    batch_id: str,
    payload: ImportBatchSceneLayerRequest,
    request: Request,
) -> dict[str, Any]:
    report = get_report_or_404(batch_id, include_targets=not use_mysql())
    return link_import_batch_scene_layer_records(report, payload, request)


@router.post("/imports/{batch_id}/publish-readiness")
def check_import_batch_publish_readiness(
    batch_id: str,
    payload: ImportBatchSceneLayerRequest,
    request: Request,
) -> dict[str, Any]:
    report = get_report_or_404(batch_id, include_targets=not use_mysql())
    return import_batch_publish_readiness(report, payload, request)


@router.get("/imports/{batch_id}/targets")
def list_import_batch_targets(
    batch_id: str,
    request: Request,
    kind: str = Query(default="blocks", pattern="^(blocks|rights)$"),
    q: str = Query(default=""),
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    report = get_report_or_404(batch_id, include_targets=not use_mysql())
    require_import_batch_summary_read_permission(report, request)
    if use_mysql():
        return list_import_batch_targets_mysql(
            batch_id,
            kind=kind,
            q=q,
            limit=limit,
            offset=offset,
        )
    return list_import_batch_targets_json(
        report,
        kind=kind,
        q=q,
        limit=limit,
        offset=offset,
    )


@router.get("/imports/{batch_id}")
def get_import_batch(
    batch_id: str,
    request: Request,
    includeTargets: bool = Query(default=False),
) -> dict[str, Any]:
    include_targets = includeTargets or not use_mysql()
    report = get_report_or_404(batch_id, include_targets=include_targets)
    if include_targets:
        require_import_batch_read_permission(report, request)
    else:
        require_import_batch_summary_read_permission(report, request)
    return public_import_report(report)


@router.get("/imports/{batch_id}/report")
def get_import_report(batch_id: str, request: Request) -> dict[str, Any]:
    report = get_report_or_404(batch_id)
    require_import_batch_read_permission(report, request)
    return report


@router.get("/imports/{batch_id}/report.json")
def download_import_report_json(batch_id: str, request: Request) -> Response:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION)
    report = get_report_or_404(batch_id)
    require_import_report_allowed(context, report)
    content = json.dumps(report, ensure_ascii=False, indent=2)
    filename = f"import-report-{batch_id}.json"
    return Response(
        content=content,
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/imports/{batch_id}/acceptance-receipt.json")
def download_import_acceptance_receipt_json(batch_id: str, request: Request) -> Response:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION)
    report = get_report_or_404(batch_id)
    require_import_report_allowed(context, report)
    append_import_batch_audit_event(
        report,
        action="export-acceptance-receipt",
        actor=context.user,
        summary={
            "permission": IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION,
            "receiptType": "import-batch-acceptance",
        },
    )
    save_import_report(report)
    filename = f"import-acceptance-receipt-{safe_download_stem(batch_id, 'batch')}.json"
    return json_download_response(filename, import_batch_acceptance_receipt(report, context))


@router.get("/imports/{batch_id}/delivery-package-receipt.json")
def download_import_delivery_package_receipt_json(batch_id: str, request: Request) -> Response:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION)
    report = get_report_or_404(batch_id)
    require_import_report_allowed(context, report)
    append_import_batch_audit_event(
        report,
        action="export-delivery-package-receipt",
        actor=context.user,
        summary={
            "permission": IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION,
            "receiptType": "import-delivery-package",
        },
    )
    save_import_report(report)
    filename = f"import-delivery-package-receipt-{safe_download_stem(batch_id, 'batch')}.json"
    return json_download_response(filename, import_delivery_package_receipt(report, context))


@router.get("/imports/{batch_id}/errors.csv")
def download_import_errors_csv(batch_id: str, request: Request) -> Response:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_EXPORT_PERMISSION)
    report = get_report_or_404(batch_id, include_targets=not use_mysql())
    require_import_report_allowed(context, report)
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(["row", "message"])
    for error in report.get("errors") or []:
        writer.writerow([error.get("row", ""), error.get("message", "")])
    content = "\ufeff" + output.getvalue()
    filename = f"import-errors-{batch_id}.csv"
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def restored_import_batch_status(report: dict[str, Any]) -> str:
    previous_status = str(report.get("statusBeforeDelete") or "").strip()
    if previous_status and previous_status != "deleted":
        return previous_status
    if report.get("rolledBackAt"):
        return "rolled_back"
    return "completed"


@router.post("/imports/{batch_id}/restore")
def restore_import_batch(batch_id: str, request: Request) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_RESTORE_PERMISSION)
    report = get_report_or_404(batch_id, include_targets=not use_mysql())
    require_import_report_allowed(context, report)
    if not report.get("deletedAt") and report.get("status") != "deleted":
        raise HTTPException(status_code=409, detail="Import batch is not deleted")
    restored_at = now_iso()
    restored_status = restored_import_batch_status(report)
    report["status"] = restored_status
    report["deletedAt"] = None
    report.pop("statusBeforeDelete", None)
    event = append_import_batch_audit_event(
        report,
        action="restore",
        actor=getattr(context, "user", ""),
        summary={"status": restored_status, "restoredAt": restored_at},
        timestamp=restored_at,
    )
    save_import_report(report)
    return {
        "ok": True,
        "restored": batch_id,
        "report": public_import_report(report),
        "event": event,
    }


@router.delete("/imports/{batch_id}")
def delete_import_batch(batch_id: str, request: Request) -> dict[str, Any]:
    context = request_context(request)
    require_permission(context, IMPORT_FOREST_BLOCKS_DELETE_PERMISSION)
    report = get_report_or_404(batch_id, include_targets=not use_mysql())
    require_import_report_allowed(context, report)
    if not report.get("deletedAt") and report.get("status") != "deleted":
        report["statusBeforeDelete"] = report.get("status") or "completed"
    deleted_at = report.get("deletedAt") or now_iso()
    report["deletedAt"] = deleted_at
    report["status"] = "deleted"
    event = append_import_batch_audit_event(
        report,
        action="delete",
        actor=getattr(context, "user", ""),
        summary={"status": "deleted", "deletedAt": deleted_at},
        timestamp=deleted_at,
    )
    save_import_report(report)
    return {"ok": True, "deleted": batch_id, "event": event}
